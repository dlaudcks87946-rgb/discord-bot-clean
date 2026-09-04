# test deploy
import discord
from discord.ext import commands, tasks
from discord import app_commands
import os
import sys
import urllib.request
import json
from flask import Flask
from threading import Thread
import asyncio
import sqlite3
import psycopg2
import time
import datetime
import calendar
import random
import re

# Railway 프로젝트에서 데이터베이스 영구 보존을 위해 PostgreSQL 서비스를 추가한 후,
# 봇 서비스의 Variables 탭에서 DATABASE_URL 변수를 추가하고 값으로 ${{Postgres.DATABASE_URL}} 을 연결해 주어야 이 환경변수를 인식합니다.
DATABASE_URL = os.getenv("DATABASE_URL")

# DATABASE_URL 환경변수(Railway PostgreSQL) 유무에 따라 자동으로 PostgreSQL 또는 로컬 SQLite 데이터베이스를 반환합니다.
def get_db_connection():
    if DATABASE_URL:
        # Railway PostgreSQL (postgres://를 postgresql://로 치환하여 psycopg2 호환)
        url = DATABASE_URL
        if url.startswith("postgres://"):
            url = url.replace("postgres://", "postgresql://", 1)
        return psycopg2.connect(url)
    else:
        # SQLite
        base_dir = os.path.dirname(os.path.abspath(__file__))
        db_path = os.path.join(base_dir, "bot_data.db")
        return sqlite3.connect(db_path)

def init_db():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    if DATABASE_URL:
        print("🔌 [데이터베이스] Railway PostgreSQL 모드로 정상 연결되었습니다.")
    else:
        print("💾 [데이터베이스] 로컬 SQLite 모드로 연결되었습니다. (Railway 배포 환경에서는 재시작 시 데이터가 유실되므로 PostgreSQL 추가가 필요합니다.)")
    
    # user_id 등 디스코드 ID를 다루기 위해 SQLite는 INTEGER, PostgreSQL은 BIGINT로 설정
    user_id_type = "BIGINT" if DATABASE_URL else "INTEGER"
    
    cursor.execute(f"""
    CREATE TABLE IF NOT EXISTS left_users (
        user_id {user_id_type} PRIMARY KEY,
        last_application TEXT,
        last_messages TEXT
    )
    """)
    
    # 마이그레이션: 기존 데이터베이스에 새로운 컬럼 안전하게 추가
    if DATABASE_URL:
        # PostgreSQL에서는 예외 발생 시 트랜잭션이 중단되므로 롤백 필수
        try:
            cursor.execute("ALTER TABLE left_users ADD COLUMN last_application TEXT")
            conn.commit()
        except Exception:
            conn.rollback()
        try:
            cursor.execute("ALTER TABLE left_users ADD COLUMN last_messages TEXT")
            conn.commit()
        except Exception:
            conn.rollback()
    else:
        # SQLite
        try:
            cursor.execute("ALTER TABLE left_users ADD COLUMN last_application TEXT")
        except Exception:
            pass
        try:
            cursor.execute("ALTER TABLE left_users ADD COLUMN last_messages TEXT")
        except Exception:
            pass
    
    cursor.execute(f"""
    CREATE TABLE IF NOT EXISTS voice_usage (
        user_id {user_id_type},
        use_date TEXT,
        seconds INTEGER DEFAULT 0,
        PRIMARY KEY (user_id, use_date)
    )
    """)
    
    cursor.execute(f"""
    CREATE TABLE IF NOT EXISTS voice_panel (
        channel_id {user_id_type} PRIMARY KEY,
        message_id {user_id_type}
    )
    """)

    cursor.execute(f"""
    CREATE TABLE IF NOT EXISTS users (
        user_id {user_id_type} PRIMARY KEY,
        xp INTEGER DEFAULT 0,
        coin INTEGER DEFAULT 0,
        random_box INTEGER DEFAULT 0,
        premium_box INTEGER DEFAULT 0,
        jackpot_box INTEGER DEFAULT 0,
        booster_until BIGINT DEFAULT 0,
        voice_minutes INTEGER DEFAULT 0
    )
    """)
    
    cursor.execute(f"""
    CREATE TABLE IF NOT EXISTS user_quests (
        user_id {user_id_type},
        quest_id TEXT,
        progress INTEGER DEFAULT 0,
        claimed INTEGER DEFAULT 0,
        quest_date TEXT,
        PRIMARY KEY (user_id, quest_id, quest_date)
    )
    """)

    cursor.execute(f"""
    CREATE TABLE IF NOT EXISTS lotto_tickets (
        id {"SERIAL" if DATABASE_URL else "INTEGER PRIMARY KEY AUTOINCREMENT"},
        user_id {user_id_type},
        round_no INTEGER,
        numbers TEXT,
        channel_id {user_id_type},
        is_checked INTEGER DEFAULT 0,
        match_count INTEGER DEFAULT 0,
        has_bonus INTEGER DEFAULT 0,
        prize_rank INTEGER DEFAULT 0,
        created_at TEXT
    )
    """)

    cursor.execute(f"""
    CREATE TABLE IF NOT EXISTS lotto_results (
        round_no INTEGER PRIMARY KEY,
        numbers TEXT,
        bonus INTEGER,
        drawn_at TEXT
    )
    """)

    # SQLite 마이그레이션 확인 (PostgreSQL에서는 신규 테이블이므로 패스)
    if not DATABASE_URL:
        try:
            cursor.execute("PRAGMA table_info(voice_usage);")
            columns = [info[1] for info in cursor.fetchall()]
            if columns and "use_date" not in columns:
                cursor.execute("DROP TABLE voice_usage;")
                print("⚠️ 마이그레이션: 기존 voice_usage 테이블을 삭제하고 새 스키마로 생성합니다.")
                cursor.execute(f"""
                CREATE TABLE IF NOT EXISTS voice_usage (
                    user_id {user_id_type},
                    use_date TEXT,
                    seconds INTEGER DEFAULT 0,
                    PRIMARY KEY (user_id, use_date)
                )
                """)
        except Exception as migration_err:
            print(f"❌ 마이그레이션 검사 중 오류 발생: {migration_err}")
            
    conn.commit()
    cursor.close()
    conn.close()

# DB 초기화 실행
init_db()


# ==========================================
# [로또 시스템 헬퍼 함수 및 클래스]
# ==========================================

def get_kst_now():
    # 서버 타임존에 무관하게 항상 한국 시간(KST, UTC+9) naive datetime 반환
    kst_tz = datetime.timezone(datetime.timedelta(hours=9))
    return datetime.datetime.now(kst_tz).replace(tzinfo=None)

def get_current_lotto_round():
    # 1회차 추첨일: 2002년 12월 7일 20:45 (토요일)
    first_round_time = datetime.datetime(2002, 12, 7, 20, 45)
    now = get_kst_now()
    delta = now - first_round_time
    weeks = delta.days // 7
    return weeks + 2

def fetch_lotto_result_sync(round_no):
    timestamp = int(time.time() * 1000)
    url = f"https://www.dhlottery.co.kr/lt645/selectPstLt645Info.do?srchLtEpsd={round_no}&_={timestamp}"
    try:
        req = urllib.request.Request(
            url, 
            headers={
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                'Referer': 'https://www.dhlottery.co.kr/'
            }
        )
        with urllib.request.urlopen(req, timeout=5) as response:
            data = json.loads(response.read().decode('utf-8'))
            lst = data.get("data", {}).get("list", [])
            if lst:
                item = lst[0]
                raw_date = str(item.get("ltRflYmd", ""))
                formatted_date = f"{raw_date[:4]}-{raw_date[4:6]}-{raw_date[6:8]}" if len(raw_date) == 8 else raw_date
                
                return {
                    "round_no": round_no,
                    "numbers": f"{item['tm1WnNo']},{item['tm2WnNo']},{item['tm3WnNo']},{item['tm4WnNo']},{item['tm5WnNo']},{item['tm6WnNo']}",
                    "bonus": item['bnsWnNo'],
                    "drawn_at": formatted_date
                }
    except Exception as e:
        print(f"❌ [로또] {round_no}회차 데이터 가져오기 실패: {e}")
    return None

async def fetch_lotto_result_async(round_no):
    loop = asyncio.get_running_loop()
    return await loop.run_in_executor(None, fetch_lotto_result_sync, round_no)

async def sync_historical_lotto_data():
    print("🔄 [로또] 최근 100회차 당첨 정보 캐싱 시작...")
    try:
        current_round = get_current_lotto_round()
        start_round = max(1, current_round - 100)
        end_round = current_round - 1
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        p = "%s" if DATABASE_URL else "?"
        cursor.execute(f"SELECT round_no FROM lotto_results WHERE round_no >= {p}", (start_round,))
        cached_rounds = {row[0] for row in cursor.fetchall()}
        
        missing_rounds = [r for r in range(start_round, end_round + 1) if r not in cached_rounds]
        
        if missing_rounds:
            print(f"🔄 [로또] 누락된 {len(missing_rounds)}개 회차 데이터 수집 중...")
            for r in missing_rounds:
                result = await fetch_lotto_result_async(r)
                if result:
                    cursor.execute(
                        f"INSERT INTO lotto_results (round_no, numbers, bonus, drawn_at) VALUES ({p}, {p}, {p}, {p})",
                        (result["round_no"], result["numbers"], result["bonus"], result["drawn_at"])
                    )
                    conn.commit()
                    await asyncio.sleep(0.1)
            print("✅ [로또] 누락된 당첨 정보 캐싱 완료!")
        else:
            print("✅ [로또] 이미 최신 당첨 정보가 캐싱되어 있습니다.")
            
        cursor.close()
        conn.close()
    except Exception as e:
        print(f"❌ [로또] 과거 데이터 캐싱 중 오류 발생: {e}")

def get_lotto_stats_from_db():
    frequencies = {i: 0 for i in range(1, 46)}
    last_seen_round = {i: 0 for i in range(1, 46)}
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        current_round = get_current_lotto_round()
        start_round = max(1, current_round - 100)
        
        p = "%s" if DATABASE_URL else "?"
        cursor.execute(
            f"SELECT round_no, numbers FROM lotto_results WHERE round_no >= {p} ORDER BY round_no ASC",
            (start_round,)
        )
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        
        actual_count = len(rows)
        for round_no, numbers_str in rows:
            nums = [int(n) for n in numbers_str.split(",")]
            for n in nums:
                if n in frequencies:
                    frequencies[n] += 1
                    last_seen_round[n] = max(last_seen_round[n], round_no)
        
        return frequencies, last_seen_round, actual_count
    except Exception as e:
        print(f"❌ [로또] 통계 데이터 계산 중 오류: {e}")
        return frequencies, last_seen_round, 0

def is_balanced_lotto(nums):
    total_sum = sum(nums)
    if not (100 <= total_sum <= 180):
        return False
    
    odds = sum(1 for n in nums if n % 2 != 0)
    if odds not in [2, 3, 4]:
        return False
        
    lows = sum(1 for n in nums if n <= 22)
    if lows not in [2, 3, 4]:
        return False
        
    return True

def has_consecutive(nums):
    for i in range(len(nums) - 1):
        if nums[i+1] - nums[i] == 1:
            return True
    return False

def generate_lotto_game(fixed_nums, excluded_nums, pattern):
    available_pool = [n for n in range(1, 46) if n not in excluded_nums and n not in fixed_nums]
    needed_count = 6 - len(fixed_nums)
    
    if needed_count < 0 or len(available_pool) < needed_count:
        return None
        
    for _ in range(1000):
        selected = random.sample(available_pool, needed_count)
        game = sorted(list(fixed_nums) + selected)
        
        if pattern == "balanced":
            if is_balanced_lotto(game):
                return game
        elif pattern == "no_consecutive":
            if not has_consecutive(game):
                return game
        elif pattern == "odd_heavy":
            odds = sum(1 for n in game if n % 2 != 0)
            if odds >= 4:
                return game
        elif pattern == "even_heavy":
            evens = sum(1 for n in game if n % 2 == 0)
            if evens >= 4:
                return game
        elif pattern == "high_heavy":
            highs = sum(1 for n in game if n >= 23)
            if highs >= 4:
                return game
        elif pattern == "low_heavy":
            lows = sum(1 for n in game if n <= 22)
            if lows >= 4:
                return game
        else:
            return game
            
    # 조건 만족 실패 시 무작위 생성
    selected = random.sample(available_pool, needed_count)
    return sorted(list(fixed_nums) + selected)

def parse_number_list(num_str):
    if not num_str:
        return set(), None
    tokens = re.split(r'[\s,]+', num_str.strip())
    numbers = set()
    for tok in tokens:
        if not tok:
            continue
        try:
            val = int(tok)
            if not (1 <= val <= 45):
                return set(), f"번호는 1에서 45 사이여야 합니다: `{tok}`"
            if val in numbers:
                return set(), f"중복된 번호가 있습니다: `{tok}`"
            numbers.add(val)
        except ValueError:
            return set(), f"올바른 숫자가 아닙니다: `{tok}`"
    return numbers, None


class LottoFilterSelect(discord.ui.Select):
    def __init__(self, current_filter):
        options = [
            discord.SelectOption(label="기본 균형형", value="balanced", description="총합, 홀짝, 고저 비율이 고루 분배된 균형 조합", emoji="⚖️"),
            discord.SelectOption(label="연속수 배제형", value="no_consecutive", description="연속되는 숫자가 없는 조합", emoji="🚫"),
            discord.SelectOption(label="홀수 강조형", value="odd_heavy", description="홀수가 4개 이상 포함된 조합", emoji="🔴"),
            discord.SelectOption(label="짝수 강조형", value="even_heavy", description="짝수가 4개 이상 포함된 조합", emoji="🔵"),
            discord.SelectOption(label="고수 강조형", value="high_heavy", description="23~45 사이 숫자가 4개 이상 포함된 조합", emoji="🟢"),
            discord.SelectOption(label="소수 강조형", value="low_heavy", description="1~22 사이 숫자가 4개 이상 포함된 조합", emoji="🟡"),
            discord.SelectOption(label="순수 무작위형", value="random", description="아무런 규칙이 적용되지 않은 완전 무작위 조합", emoji="🎰")
        ]
        for opt in options:
            if opt.value == current_filter:
                opt.default = True
                break
        super().__init__(placeholder="적용할 AI 패턴 필터를 선택하세요...", min_values=1, max_values=1, options=options, custom_id="lotto_filter_select")

    async def callback(self, interaction: discord.Interaction):
        if interaction.user.id != self.view.user_id:
            await interaction.response.send_message("❌ 이 번호 생성기의 소유자만 필터를 변경할 수 있습니다.", ephemeral=True)
            return
            
        self.view.pattern = self.values[0]
        self.view.regenerate_numbers()
        
        for opt in self.options:
            opt.default = (opt.value == self.view.pattern)
            
        embed = self.view.create_embed()
        await interaction.response.edit_message(embed=embed, view=self.view)


class LottoRecommendView(discord.ui.View):
    def __init__(self, user_id, count, fixed_nums, excluded_nums, pattern="balanced"):
        super().__init__(timeout=180)
        self.user_id = user_id
        self.count = count
        self.fixed_nums = fixed_nums
        self.excluded_nums = excluded_nums
        self.pattern = pattern
        self.games = []
        
        self.add_item(LottoFilterSelect(self.pattern))
        self.regenerate_numbers()

    def regenerate_numbers(self):
        self.games = []
        for _ in range(self.count):
            game = generate_lotto_game(self.fixed_nums, self.excluded_nums, self.pattern)
            if game:
                self.games.append(game)
            else:
                available = [n for n in range(1, 46) if n not in self.excluded_nums and n not in self.fixed_nums]
                self.games.append(sorted(list(self.fixed_nums) + random.sample(available, 6 - len(self.fixed_nums))))

    def create_embed(self):
        frequencies, last_seen, actual_count = get_lotto_stats_from_db()
        current_round = get_current_lotto_round()
        
        filter_names = {
            "balanced": "⚖️ 기본 균형형",
            "no_consecutive": "🚫 연속수 배제형",
            "odd_heavy": "🔴 홀수 강조형",
            "even_heavy": "🔵 짝수 강조형",
            "high_heavy": "🟢 고수 강조형",
            "low_heavy": "🟡 소수 강조형",
            "random": "🎰 순수 무작위형"
        }
        
        embed = discord.Embed(
            title="🎰 HEAVEN AI 로또 연구소",
            description=(
                f"원하는 패턴 필터를 적용하여 추천 번호를 생성해 보세요!\n"
                f"**💾 번호 저장**을 누르면 추첨일(매주 토요일 저녁 8시 45분) 결과 발표 후 **자동으로 당첨 여부를 정산해 DM/채널로 알림**을 보내드립니다!\n\n"
                f"ℹ️ **현재 적용된 AI 필터:** `{filter_names[self.pattern]}`\n"
                f"ℹ️ **다음 추첨 회차:** `제 {current_round}회차 대비`"
            ),
            color=0xFF007F
        )
        
        def get_emoji(num):
            if 1 <= num <= 10: return "🟡"
            elif 11 <= num <= 20: return "🔵"
            elif 21 <= num <= 30: return "🔴"
            elif 31 <= num <= 40: return "⚫"
            else: return "🟢"

        if self.count == 1:
            game = self.games[0]
            formatted_numbers = "  ".join([f"{get_emoji(num)} `{num:02d}`" for num in game])
            embed.add_field(name="🎫 추천 번호 조합", value=formatted_numbers, inline=False)
            
            total_sum = sum(game)
            odds = sum(1 for n in game if n % 2 != 0)
            evens = 6 - odds
            lows = sum(1 for n in game if n <= 22)
            highs = 6 - lows
            
            diffs = set()
            for i in range(len(game)):
                for j in range(i + 1, len(game)):
                    diffs.add(game[j] - game[i])
            ac_val = len(diffs) - (6 - 1)
            
            consec_pairs = []
            for i in range(len(game) - 1):
                if game[i+1] - game[i] == 1:
                    consec_pairs.append(f"({game[i]}, {game[i+1]})")
            consec_text = ", ".join(consec_pairs) if consec_pairs else "없음"
            
            freq_details = []
            cold_details = []
            for num in game:
                freq = frequencies.get(num, 0)
                last_rd = last_seen.get(num, 0)
                cold_weeks = current_round - last_rd if last_rd > 0 else 100
                cold_text = f"{cold_weeks}주" if cold_weeks < 100 else "100주+"
                
                freq_details.append(f"`{num:02d}`({freq}회)")
                cold_details.append(f"`{num:02d}`({cold_text})")
            
            avg_freq = sum(frequencies.get(num, 0) for num in game) / 6
            avg_cold = sum((current_round - last_seen.get(num, 0)) if last_seen.get(num, 0) > 0 else 100 for num in game) / 6
            
            analysis_text = (
                f"▪️ **총합:** `{total_sum}` {'(균형: 100~180)' if 100 <= total_sum <= 180 else '(비균형)'}\n"
                f"▪️ **홀짝 비율:** `{odds}:{evens}`\n"
                f"▪️ **고저 비율:** `{lows}:{highs}` (Low: 1~22, High: 23~45)\n"
                f"▪️ **산술 복잡도 (AC값):** `{ac_val}` (5 이상 권장)\n"
                f"▪️ **연속 번호 쌍:** `{consec_text}`\n"
                f"▪️ **평균 출현 빈도:** `{avg_freq:.1f}회` (11~15회 권장, 최근 100회 기준)\n"
                f"▪️ **평균 미출현 기간:** `{avg_cold:.1f}주` (6~10주 권장)"
            )
            embed.add_field(name="📊 실시간 패턴 및 통계 분석", value=analysis_text, inline=False)
            
            detail_stat_text = (
                f"▪️ **번호별 출현 빈도:** {', '.join(freq_details)}\n"
                f"▪️ **번호별 미출현 기간:** {', '.join(cold_details)}"
            )
            embed.add_field(name="🔍 최근 100회차 심층 통계", value=detail_stat_text, inline=False)
            
        else:
            game_list_text = []
            for i, game in enumerate(self.games, 1):
                formatted_numbers = " ".join([f"{get_emoji(num)} `{num:02d}`" for num in game])
                
                total_sum = sum(game)
                odds = sum(1 for n in game if n % 2 != 0)
                evens = 6 - odds
                lows = sum(1 for n in game if n <= 22)
                highs = 6 - lows
                
                mini_stat = f"└ `합:{total_sum:03d} | 홀짝 {odds}:{evens} | 고저 {lows}:{highs}`"
                game_list_text.append(f"**🎫 {i:02d}번째 게임**\n{formatted_numbers}\n{mini_stat}")
                
            chunk_size = 5
            for idx in range(0, len(game_list_text), chunk_size):
                chunk = game_list_text[idx:idx+chunk_size]
                embed.add_field(
                    name=f"📋 추천 조합 목록 ({idx+1}~{min(self.count, idx+chunk_size)}번째)",
                    value="\n".join(chunk),
                    inline=False
                )
                
            all_sums = [sum(g) for g in self.games]
            avg_sum = sum(all_sums) / len(self.games)
            embed.set_footer(text=f"전체 {self.count}개 게임 평균 총합: {avg_sum:.1f} | 추첨 당첨 보장 없음")

        rule_texts = []
        if self.fixed_nums:
            rule_texts.append(f"고정수: `{' '.join(f'{n:02d}' for n in sorted(list(self.fixed_nums)))}`")
        if self.excluded_nums:
            rule_texts.append(f"제외수: `{' '.join(f'{n:02d}' for n in sorted(list(self.excluded_nums)))}`")
        if rule_texts:
            embed.add_field(name="⚙️ 커스텀 규칙 적용", value=" | ".join(rule_texts), inline=False)
            
        return embed

    @discord.ui.button(label="재생성 (Re-roll)", style=discord.ButtonStyle.secondary, emoji="🔄", custom_id="lotto_reroll_btn", row=1)
    async def reroll(self, interaction: discord.Interaction, button: discord.ui.Button):
        if interaction.user.id != self.user_id:
            await interaction.response.send_message("❌ 이 번호 생성기의 소유자만 조작할 수 있습니다.", ephemeral=True)
            return
            
        self.regenerate_numbers()
        
        # 다시 저장할 수 있도록 저장 버튼 리셋
        for child in self.children:
            if getattr(child, "custom_id", None) == "lotto_save_btn":
                child.disabled = False
                child.label = "번호 저장 (Save)"
                child.style = discord.ButtonStyle.success
                break
                
        embed = self.create_embed()
        await interaction.response.edit_message(embed=embed, view=self)

    @discord.ui.button(label="번호 저장 (Save)", style=discord.ButtonStyle.success, emoji="💾", custom_id="lotto_save_btn", row=1)
    async def save_tickets(self, interaction: discord.Interaction, button: discord.ui.Button):
        if interaction.user.id != self.user_id:
            await interaction.response.send_message("❌ 이 번호 생성기의 소유자만 번호를 저장할 수 있습니다.", ephemeral=True)
            return
            
        current_round = get_current_lotto_round()
        now_str = get_kst_now().strftime("%Y-%m-%d %H:%M:%S")
        
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            p = "%s" if DATABASE_URL else "?"
            
            for game in self.games:
                nums_str = ",".join(str(n) for n in game)
                cursor.execute(
                    f"INSERT INTO lotto_tickets (user_id, round_no, numbers, channel_id, is_checked, match_count, has_bonus, prize_rank, created_at) VALUES ({p}, {p}, {p}, {p}, 0, 0, 0, 0, {p})",
                    (interaction.user.id, current_round, nums_str, interaction.channel_id, now_str)
                )
            conn.commit()
            cursor.close()
            conn.close()
            
            button.disabled = True
            button.label = "저장 완료 (Saved)"
            button.style = discord.ButtonStyle.secondary
            
            await interaction.response.edit_message(view=self)
            
            success_embed = discord.Embed(
                title="💾 로또 번호 저장 완료!",
                description=(
                    f"✅ **총 {len(self.games)}개 게임**이 성공적으로 저장되었습니다.\n"
                    f"📅 **대상 회차:** 제 {current_round}회차 추첨 대비\n"
                    f"🔔 **자동 정산:** 추첨이 완료되는 토요일 밤 8시 45분 이후 자동으로 결과를 채널 및 개인 DM으로 알려드립니다."
                ),
                color=0x39FF14
            )
            await interaction.followup.send(embed=success_embed, ephemeral=True)
            
        except Exception as e:
            print(f"❌ [로또] 번호 저장 중 오류 발생: {e}")
            await interaction.response.send_message("❌ 번호 저장 중 데이터베이스 오류가 발생했습니다. 다시 시도해 주세요.", ephemeral=True)


@tasks.loop(minutes=10.0)
async def lotto_check_loop():
    now_kst = get_kst_now()
    # 토요일 (weekday 5) 이고, 20:45 ~ 21:30 KST 사이일 때 결과 조회 시도
    if now_kst.weekday() == 5 and (20, 45) <= (now_kst.hour, now_kst.minute) <= (21, 30):
        current_round = get_current_lotto_round()
        target_round = current_round - 1
        
        conn = get_db_connection()
        cursor = conn.cursor()
        p = "%s" if DATABASE_URL else "?"
        cursor.execute(f"SELECT round_no FROM lotto_results WHERE round_no = {p}", (target_round,))
        cached = cursor.fetchone()
        
        if not cached:
            print(f"📡 [로또] {target_round}회차 추첨 결과 조회 시도 중...")
            result = await fetch_lotto_result_async(target_round)
            if result:
                cursor.execute(
                    f"INSERT INTO lotto_results (round_no, numbers, bonus, drawn_at) VALUES ({p}, {p}, {p}, {p})",
                    (result["round_no"], result["numbers"], result["bonus"], result["drawn_at"])
                )
                conn.commit()
                print(f"✅ [로또] {target_round}회차 당첨 번호 저장 성공: {result['numbers']} + {result['bonus']}")
                
                await check_and_notify_lotto_tickets(target_round, result["numbers"], result["bonus"])
            else:
                print(f"⚠️ [로또] {target_round}회차 추첨 결과를 아직 가져올 수 없습니다. 10분 후 재시도합니다.")
        
        cursor.close()
        conn.close()

async def check_and_notify_lotto_tickets(round_no, win_numbers_str, bonus):
    win_nums = set(int(n) for n in win_numbers_str.split(","))
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        p = "%s" if DATABASE_URL else "?"
        
        cursor.execute(
            f"SELECT id, user_id, numbers, channel_id FROM lotto_tickets WHERE round_no = {p} AND is_checked = 0",
            (round_no,)
        )
        tickets = cursor.fetchall()
        
        if not tickets:
            cursor.close()
            conn.close()
            return
            
        print(f"📊 [로또] {round_no}회차 티켓 정산 시작 (대상: {len(tickets)}개)")
        
        for ticket_id, user_id, numbers_str, channel_id in tickets:
            user_nums = [int(n) for n in numbers_str.split(",")]
            
            matches = set(user_nums).intersection(win_nums)
            match_count = len(matches)
            has_bonus = 1 if bonus in user_nums else 0
            
            prize_rank = 0
            if match_count == 6:
                prize_rank = 1
            elif match_count == 5 and has_bonus == 1:
                prize_rank = 2
            elif match_count == 5:
                prize_rank = 3
            elif match_count == 4:
                prize_rank = 4
            elif match_count == 3:
                prize_rank = 5
                
            cursor.execute(
                f"UPDATE lotto_tickets SET is_checked = 1, match_count = {p}, has_bonus = {p}, prize_rank = {p} WHERE id = {p}",
                (match_count, has_bonus, prize_rank, ticket_id)
            )
            conn.commit()
            
            await send_lotto_notification(user_id, round_no, user_nums, win_nums, bonus, match_count, prize_rank, channel_id)
            
        cursor.close()
        conn.close()
    except Exception as e:
        print(f"❌ [로또] 티켓 정산 처리 중 오류 발생: {e}")

async def send_lotto_notification(user_id, round_no, user_nums, win_nums, bonus, match_count, prize_rank, channel_id):
    rank_info = {
        1: ("🥇 1등 (6개 일치)", "당첨금: 동행복권 공식 홈페이지 참조"),
        2: ("🥈 2등 (5개 + 보너스 일치)", "당첨금: 동행복권 공식 홈페이지 참조"),
        3: ("🥉 3등 (5개 일치)", "당첨금: 동행복권 공식 홈페이지 참조"),
        4: ("💎 4등 (4개 일치)", "당첨금: 50,000원"),
        5: ("🍀 5등 (3개 일치)", "당첨금: 5,000원"),
        0: ("❌ 낙첨", "다음 기회에 다시 도전해보세요!")
    }
    
    rank_name, prize_desc = rank_info[prize_rank]
    
    def get_emoji(num):
        if 1 <= num <= 10: return "🟡"
        elif 11 <= num <= 20: return "🔵"
        elif 21 <= num <= 30: return "🔴"
        elif 31 <= num <= 40: return "⚫"
        else: return "🟢"
        
    formatted_user_nums = " ".join([f"{get_emoji(n)} `{n:02d}`" for n in user_nums])
    sorted_win_list = sorted(list(win_nums))
    formatted_win_nums = " ".join([f"{get_emoji(n)} `{n:02d}`" for n in sorted_win_list]) + f"  ➕  {get_emoji(bonus)} `{bonus:02d}` (보너스)"
    
    embed_color = 0x39FF14 if prize_rank > 0 else 0xFF0055
    
    embed = discord.Embed(
        title=f"🎰 제 {round_no}회 로또 추첨 결과 알림",
        description=f"<@{user_id}> 님이 등록하신 티켓의 정산 결과입니다.",
        color=embed_color
    )
    
    embed.add_field(name="🗳️ 공식 당첨 번호", value=formatted_win_nums, inline=False)
    embed.add_field(name="🎫 내 추천 번호", value=formatted_user_nums, inline=False)
    embed.add_field(name="📊 분석 결과", value=f"**일치 개수:** {match_count}개 일치 (보너스 일치: {'예' if bonus in user_nums else '아니오'})\n**최종 결과: {rank_name}**\n*{prize_desc}*", inline=False)
    embed.set_footer(text="HEAVEN AI 로또 연구소")
    
    user = bot.get_user(user_id)
    if not user:
        try:
            user = await bot.fetch_user(user_id)
        except Exception:
            pass
            
    if user:
        try:
            await user.send(embed=embed)
        except Exception as e:
            print(f"❌ [로또] <@{user_id}>에게 개인 DM 전송 실패: {e}")
            
    # 로또 당첨/낙첨 결과는 개인의 사생활 보호를 위해 채널에 전송하지 않고 개인 DM으로만 전송합니다.
    # (기존 채널 알림 코드는 본인만 볼 수 있도록 제거됨)


# Active voice sessions tracking (user_id -> join_timestamp)
active_sessions = {}

def get_current_date():
    # 현재 시간에서 6시간을 빼서 오전 6시를 하루의 시작 기준으로 설정
    adjusted_now = datetime.datetime.now() - datetime.timedelta(hours=6)
    return adjusted_now.strftime("%Y-%m-%d")

def get_recent_dates():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT use_date FROM voice_usage ORDER BY use_date DESC LIMIT 100")
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        return [row[0] for row in rows]
    except Exception as e:
        print(f"❌ 날짜 목록 조회 중 오류 발생: {e}")
        return []

def add_voice_time(user_id, use_date, seconds):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        p = "%s" if DATABASE_URL else "?"
        # ON CONFLICT 구문은 SQLite 3.24.0+ 및 PostgreSQL에서 작동합니다.
        query = f"""
        INSERT INTO voice_usage (user_id, use_date, seconds)
        VALUES ({p}, {p}, {p})
        ON CONFLICT (user_id, use_date)
        DO UPDATE SET seconds = voice_usage.seconds + EXCLUDED.seconds
        """
        cursor.execute(query, (user_id, use_date, seconds))
        conn.commit()
        cursor.close()
        conn.close()
    except Exception as e:
        print(f"❌ DB 이용시간 저장 중 오류 발생: {e}")

def get_realtime_today_stats():
    try:
        today_str = get_current_date()
        conn = get_db_connection()
        cursor = conn.cursor()
        p = "%s" if DATABASE_URL else "?"
        cursor.execute(f"SELECT user_id, seconds FROM voice_usage WHERE use_date = {p}", (today_str,))
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        
        stats = {user_id: seconds for user_id, seconds in rows}
        
        # 현재 음성 채널에 접속 중인 멤버들의 세션 시간 실시간 합산
        now = time.time()
        for uid, join_time in active_sessions.items():
            active_duration = int(now - join_time)
            if active_duration > 0:
                stats[uid] = stats.get(uid, 0) + active_duration
                
        sorted_stats = sorted(stats.items(), key=lambda x: x[1], reverse=True)
        return sorted_stats
    except Exception as e:
        print(f"❌ 실시간 오늘의 통계 집계 중 오류 발생: {e}")
        return []

def get_realtime_cumulative_stats():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT user_id, SUM(seconds) FROM voice_usage GROUP BY user_id")
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        
        stats = {user_id: (int(secs) if secs is not None else 0) for user_id, secs in rows}
        
        # 현재 음성 채널에 접속 중인 멤버들의 세션 시간 실시간 합산
        now = time.time()
        for uid, join_time in active_sessions.items():
            active_duration = int(now - join_time)
            if active_duration > 0:
                stats[uid] = stats.get(uid, 0) + active_duration
                
        sorted_stats = sorted(
            [(uid, secs) for uid, secs in stats.items() if secs > 0],
            key=lambda x: x[1],
            reverse=True
        )
        return sorted_stats
    except Exception as e:
        print(f"❌ 실시간 누적 통계 집계 중 오류 발생: {e}")
        return []

def format_time(seconds):
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    secs = seconds % 60
    parts = []
    if hours > 0:
        parts.append(f"{hours}시간")
    if minutes > 0:
        parts.append(f"{minutes}분")
    if secs > 0 or not parts:
        parts.append(f"{secs}초")
    return " ".join(parts)

# ==========================================
# [월간 활동 기록 및 미접속자 추적 헬퍼 함수]
# ==========================================

def get_month_date_range(year=None, month=None):
    """특정 연월의 시작일(YYYY-MM-01)과 말일(YYYY-MM-LastDay) 반환"""
    now = get_kst_now()
    if year is None:
        year = now.year
    if month is None:
        month = now.month
    
    _, last_day = calendar.monthrange(year, month)
    start_date = f"{year:04d}-{month:02d}-01"
    end_date = f"{year:04d}-{month:02d}-{last_day:02d}"
    return start_date, end_date, year, month, last_day

def get_monthly_voice_stats(year=None, month=None):
    """특정 연월(1일~말일) 음성 이용 시간 통계 및 랭킹 집계"""
    start_date, end_date, y, m, last_day = get_month_date_range(year, month)
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        p = "%s" if DATABASE_URL else "?"
        query = f"""
        SELECT user_id, SUM(seconds) 
        FROM voice_usage 
        WHERE use_date >= {p} AND use_date <= {p} 
        GROUP BY user_id
        """
        cursor.execute(query, (start_date, end_date))
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        
        stats = {user_id: (int(secs) if secs is not None else 0) for user_id, secs in rows}
        
        # 현재 월인 경우 실시간 접속 세션 반영
        now_kst = get_kst_now()
        if y == now_kst.year and m == now_kst.month:
            now_ts = time.time()
            for uid, join_time in active_sessions.items():
                active_duration = int(now_ts - join_time)
                if active_duration > 0:
                    stats[uid] = stats.get(uid, 0) + active_duration
                    
        sorted_stats = sorted(
            [(uid, secs) for uid, secs in stats.items() if secs > 0],
            key=lambda x: x[1],
            reverse=True
        )
        total_seconds = sum(secs for _, secs in sorted_stats)
        return {
            "year": y,
            "month": m,
            "start_date": start_date,
            "end_date": end_date,
            "last_day": last_day,
            "rankings": sorted_stats,
            "total_users": len(sorted_stats),
            "total_seconds": total_seconds
        }
    except Exception as e:
        print(f"❌ 월간 통계 집계 오류: {e}")
        return {
            "year": y,
            "month": m,
            "start_date": start_date,
            "end_date": end_date,
            "last_day": last_day,
            "rankings": [],
            "total_users": 0,
            "total_seconds": 0
        }

def get_available_months():
    """DB에 기록된 음성 데이터가 존재하는 연월(YYYY-MM) 목록 반환 (최신순)"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT SUBSTR(use_date, 1, 7) FROM voice_usage ORDER BY 1 DESC")
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        months = [r[0] for r in rows if r[0]]
        current_ym = get_kst_now().strftime("%Y-%m")
        if current_ym not in months:
            months.insert(0, current_ym)
        return months[:12]
    except Exception as e:
        print(f"❌ 가용 월 목록 조회 오류: {e}")
        return [get_kst_now().strftime("%Y-%m")]

def get_inactive_members(guild: discord.Guild, days_threshold: int = 14):
    """
    서버 멤버 중 지정일수(기본 14일/2주) 이상 음성 채널에 접속하지 않은 멤버 목록을
    미접속 일수 내림차순(오래 안 들어온 순)으로 정렬하여 반환합니다.
    """
    if not guild:
        return []
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT user_id, MAX(use_date) FROM voice_usage GROUP BY user_id")
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        
        last_seen_map = {row[0]: row[1] for row in rows if row[1]}
        
        now_kst = get_kst_now()
        today_date = now_kst.date()
        kst_tz = datetime.timezone(datetime.timedelta(hours=9))
        
        inactive_list = []
        
        for member in guild.members:
            if member.bot:
                continue
            
            # 현재 음성 채널에 접속 중이면 미접속 0일
            if member.id in active_sessions or (member.voice and member.voice.channel):
                continue
            
            last_use_str = last_seen_map.get(member.id)
            if last_use_str:
                try:
                    last_date = datetime.datetime.strptime(last_use_str, "%Y-%m-%d").date()
                    inactive_days = (today_date - last_date).days
                    has_record = True
                except Exception:
                    inactive_days = 0
                    has_record = False
            else:
                # 음성 기록이 전혀 없는 유저는 서버 가입일 기준 계산
                if member.joined_at:
                    join_kst = member.joined_at.astimezone(kst_tz).date()
                    inactive_days = (today_date - join_kst).days
                    last_use_str = join_kst.strftime("%Y-%m-%d")
                else:
                    inactive_days = 999
                    last_use_str = "기록 없음"
                has_record = False
            
            if inactive_days >= days_threshold:
                inactive_list.append({
                    "member": member,
                    "user_id": member.id,
                    "inactive_days": inactive_days,
                    "last_date_str": last_use_str,
                    "has_record": has_record,
                    "display_name": member.display_name
                })
                
        # 미접속 일수 큰 순서대로 정렬
        inactive_list.sort(key=lambda x: x["inactive_days"], reverse=True)
        return inactive_list
    except Exception as e:
        print(f"❌ 미접속자 집계 오류: {e}")
        return []


# ==========================================
# HEAVEN 시즌 패스 DB 헬퍼 및 비즈니스 로직
# ==========================================
def ensure_user(user_id: int):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        if DATABASE_URL:
            cursor.execute("INSERT INTO users(user_id) VALUES (%s) ON CONFLICT (user_id) DO NOTHING", (user_id,))
        else:
            cursor.execute("INSERT OR IGNORE INTO users(user_id) VALUES (?)", (user_id,))
        conn.commit()
        cursor.close()
        conn.close()
    except Exception as e:
        print(f"❌ ensure_user 오류: {e}")

def get_user(user_id: int):
    ensure_user(user_id)
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        p = "%s" if DATABASE_URL else "?"
        cursor.execute(f"""
            SELECT xp, coin, random_box, premium_box, jackpot_box, booster_until, voice_minutes
            FROM users WHERE user_id={p}
        """, (user_id,))
        row = cursor.fetchone()
        cursor.close()
        conn.close()
        return row
    except Exception as e:
        print(f"❌ get_user 오류: {e}")
        return (0, 0, 0, 0, 0, 0, 0)

def add_coin(user_id: int, amount: int):
    ensure_user(user_id)
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        p = "%s" if DATABASE_URL else "?"
        cursor.execute(f"UPDATE users SET coin = coin + {p} WHERE user_id={p}", (amount, user_id))
        conn.commit()
        cursor.close()
        conn.close()
    except Exception as e:
        print(f"❌ add_coin 오류: {e}")

def add_item(user_id: int, item: str, amount: int):
    allowed_items = ["random_box", "premium_box", "jackpot_box"]
    if item not in allowed_items:
        raise ValueError("Invalid item name")
    ensure_user(user_id)
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        p = "%s" if DATABASE_URL else "?"
        cursor.execute(f"UPDATE users SET {item} = {item} + {p} WHERE user_id={p}", (amount, user_id))
        conn.commit()
        cursor.close()
        conn.close()
    except Exception as e:
        print(f"❌ add_item 오류: {e}")

def use_item(user_id: int, item: str):
    allowed_items = ["random_box", "premium_box", "jackpot_box"]
    if item not in allowed_items:
        raise ValueError("Invalid item name")
    ensure_user(user_id)
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        p = "%s" if DATABASE_URL else "?"
        cursor.execute(f"SELECT {item} FROM users WHERE user_id={p}", (user_id,))
        row = cursor.fetchone()
        count = row[0] if row else 0

        if count <= 0:
            cursor.close()
            conn.close()
            return False

        cursor.execute(f"UPDATE users SET {item} = {item} - 1 WHERE user_id={p}", (user_id,))
        conn.commit()
        cursor.close()
        conn.close()
        return True
    except Exception as e:
        print(f"❌ use_item 오류: {e}")
        return False

# 레벨업 보상 테이블
REWARDS = {
    5: ("coin", None, 500, "💰 재화 500"),
    10: ("item", "random_box", 1, "📦 랜덤 상자 1개"),
    15: ("coin", None, 1000, "💰 재화 1,000"),
    20: ("item", "random_box", 2, "📦 랜덤 상자 2개"),
    25: ("coin", None, 2500, "💰 재화 2,500"),
    30: ("item", "premium_box", 1, "🎁 프리미엄 상자 1개"),
    35: ("coin", None, 3000, "💰 재화 3,000"),
    40: ("item", "premium_box", 2, "🎁 프리미엄 상자 2개"),
    45: ("item", "random_box", 5, "📦 랜덤 상자 5개"),
    50: ("item", "jackpot_box", 1, "👑 잭팟 상자 1개")
}

DAILY_QUESTS = {
    "voice_30m": {
        "title": "🎙️ 음성 채널 30분 참여하기",
        "target": 30,
        "xp_reward": 100,
        "coin_reward": 500
    },
    "open_box": {
        "title": "📦 아무 상자 1회 오픈하기",
        "target": 1,
        "xp_reward": 50,
        "coin_reward": 300
    },
    "buy_shop": {
        "title": "🛒 상점에서 상품 1회 구매하기",
        "target": 1,
        "xp_reward": 50,
        "coin_reward": 300
    }
}

def level_from_xp(xp: int):
    level = 1
    need = 300
    while xp >= need:
        xp -= need
        level += 1
        need = 300 + (level - 1) * 100
    return level, xp, need

def progress_bar(current, total, size=10):
    filled = int((current / total) * size) if total > 0 else 0
    return "█" * filled + "░" * (size - filled)

def next_reward(level: int):
    for lv, (_, _, _, r_name) in REWARDS.items():
        if lv > level:
            return f"Lv.{lv} 달성 시 {r_name}"
    return "모든 패스 보상 달성 완료"

def get_season_pass_rankings():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT user_id, xp FROM users ORDER BY xp DESC")
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        return rows
    except Exception as e:
        print(f"❌ get_season_pass_rankings 오류: {e}")
        return []

def check_and_grant_level_rewards(cursor, p, user_id, old_xp, new_xp):
    old_level, _, _ = level_from_xp(old_xp)
    new_level, _, _ = level_from_xp(new_xp)
    rewards_granted = []
    if new_level > old_level:
        for lv in range(old_level + 1, new_level + 1):
            if lv in REWARDS:
                r_type, r_target, r_amount, r_name = REWARDS[lv]
                if r_type == "coin":
                    cursor.execute(f"UPDATE users SET coin = coin + {p} WHERE user_id={p}", (r_amount, user_id))
                elif r_type == "item":
                    cursor.execute(f"UPDATE users SET {r_target} = {r_target} + {p} WHERE user_id={p}", (r_amount, user_id))
                elif r_type == "booster":
                    now = int(time.time())
                    duration = r_amount * 86400
                    cursor.execute(f"SELECT booster_until FROM users WHERE user_id={p}", (user_id,))
                    row = cursor.fetchone()
                    curr_booster = row[0] if row else 0
                    new_booster = max(curr_booster, now) + duration
                    cursor.execute(f"UPDATE users SET booster_until = {p} WHERE user_id={p}", (new_booster, user_id))
                rewards_granted.append(r_name)
    return rewards_granted
def update_quest_progress(user_id: int, quest_id: str, amount: int = 1):
    try:
        today = get_current_date()
        ensure_user(user_id)
        conn = get_db_connection()
        cursor = conn.cursor()
        p = "%s" if DATABASE_URL else "?"
        
        if DATABASE_URL:
            cursor.execute("""
                INSERT INTO user_quests (user_id, quest_id, progress, claimed, quest_date)
                VALUES (%s, %s, %s, 0, %s)
                ON CONFLICT (user_id, quest_id, quest_date)
                DO UPDATE SET progress = user_quests.progress + EXCLUDED.progress
            """, (user_id, quest_id, amount, today))
        else:
            cursor.execute("""
                INSERT INTO user_quests (user_id, quest_id, progress, claimed, quest_date)
                VALUES (?, ?, ?, 0, ?)
                ON CONFLICT (user_id, quest_id, quest_date)
                DO UPDATE SET progress = user_quests.progress + EXCLUDED.progress
            """, (user_id, quest_id, amount, today))
            
        conn.commit()
        cursor.close()
        conn.close()
    except Exception as e:
        print(f"❌ update_quest_progress 오류: {e}")

def claim_all_quests_calc(user_id: int):
    today = get_current_date()
    ensure_user(user_id)
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        p = "%s" if DATABASE_URL else "?"
        
        cursor.execute(f"""
            SELECT quest_id, progress, claimed FROM user_quests
            WHERE user_id = {p} AND quest_date = {p}
        """, (user_id, today))
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        
        quest_data = {row[0]: {"progress": row[1], "claimed": row[2]} for row in rows}
        
        total_xp = 0
        total_coins = 0
        claimed_quests = []
        
        for q_id, q_info in DAILY_QUESTS.items():
            db_info = quest_data.get(q_id, {"progress": 0, "claimed": 0})
            if db_info["progress"] >= q_info["target"] and not db_info["claimed"]:
                total_xp += q_info["xp_reward"]
                total_coins += q_info["coin_reward"]
                claimed_quests.append(q_id)
                
        if not claimed_quests:
            return False, "수령할 수 있는 퀘스트 보상이 없습니다."
            
        return True, (total_xp, total_coins, claimed_quests)
        
    except Exception as e:
        print(f"❌ claim_all_quests_calc 오류: {e}")
        return False, "보상 계산 중 오류가 발생했습니다."

def apply_claimed_quests(user_id: int, total_xp: int, total_coins: int, claimed_quests: list):
    today = get_current_date()
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        p = "%s" if DATABASE_URL else "?"
        
        placeholders = ", ".join([p] * len(claimed_quests))
        cursor.execute(f"""
            UPDATE user_quests SET claimed = 1
            WHERE user_id = {p} AND quest_date = {p} AND quest_id IN ({placeholders})
        """, (user_id, today) + tuple(claimed_quests))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        add_coin(user_id, total_coins)
        rewards_granted = add_xp(user_id, total_xp)
        
        msg = f"🎉 **일일 퀘스트 보상 일괄 수령 완료**\n\n계정으로 아래 보상이 즉시 지급되었습니다:\n\n* 💰 **+{total_coins:,} 코인**\n* ⭐ **+{total_xp:,} XP**"
        if rewards_granted:
            msg += f"\n\n🎁 **레벨업 달성 보상 획득!**\n└ {', '.join(rewards_granted)}"
            
        return msg
    except Exception as e:
        print(f"❌ apply_claimed_quests 오류: {e}")
        return "보상을 지급하는 도중 오류가 발생했습니다."

def add_xp(user_id: int, amount: int):
    ensure_user(user_id)
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        p = "%s" if DATABASE_URL else "?"
        
        cursor.execute(f"SELECT xp FROM users WHERE user_id={p}", (user_id,))
        row = cursor.fetchone()
        old_xp = row[0] if row else 0
        new_xp = old_xp + amount
        
        cursor.execute(f"UPDATE users SET xp = xp + {p} WHERE user_id={p}",
                       (amount, user_id))
        
        rewards = check_and_grant_level_rewards(cursor, p, user_id, old_xp, new_xp)
        conn.commit()
        cursor.close()
        conn.close()
        return rewards
    except Exception as e:
        print(f"❌ add_xp 오류: {e}")
        return []

# 상점 구매 비즈니스 로직
def buy_shop_item(user_id: int, item_type: str, cost: int, count: int = 1):
    ensure_user(user_id)
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        p = "%s" if DATABASE_URL else "?"
        cursor.execute(f"SELECT coin, booster_until FROM users WHERE user_id={p}", (user_id,))
        row = cursor.fetchone()
        if not row:
            cursor.close()
            conn.close()
            return False, "유저 정보를 찾을 수 없습니다."
        
        coins, booster_until = row
        total_cost = cost * count
        if coins < total_cost:
            cursor.close()
            conn.close()
            return False, f"❌ 재화가 부족합니다. (보유: {coins:,} / 필요: {total_cost:,})"
        
        cursor.execute(f"UPDATE users SET coin = coin - {p} WHERE user_id={p}", (total_cost, user_id))
        
        now = int(time.time())
        if item_type == "random_box":
            cursor.execute(f"UPDATE users SET random_box = random_box + {p} WHERE user_id={p}", (count, user_id))
            msg = f"📦 랜덤 상자 {count}개를 구매했습니다!"
        elif item_type == "premium_box":
            cursor.execute(f"UPDATE users SET premium_box = premium_box + {p} WHERE user_id={p}", (count, user_id))
            msg = f"🎁 프리미엄 랜덤 상자 {count}개를 구매했습니다!"
        elif item_type == "booster_1d":
            new_booster = max(booster_until, now) + count * 86400
            cursor.execute(f"UPDATE users SET booster_until = {p} WHERE user_id={p}", (new_booster, user_id))
            msg = f"💎 XP 부스터 1일 {count}개를 구매했습니다!"
        elif item_type == "booster_7d":
            new_booster = max(booster_until, now) + count * 7 * 86400
            cursor.execute(f"UPDATE users SET booster_until = {p} WHERE user_id={p}", (new_booster, user_id))
            msg = f"💎 XP 부스터 7일 {count}개를 구매했습니다!"
        else:
            cursor.close()
            conn.close()
            return False, "올바르지 않은 상품입니다."
            
        conn.commit()
        cursor.close()
        conn.close()
        update_quest_progress(user_id, "buy_shop", count)
        return True, msg
    except Exception as e:
        print(f"❌ buy_shop_item 오류: {e}")
        return False, "구매 처리 중 오류가 발생했습니다."

# 상자 열기 비즈니스 로직
def open_random_box(user_id: int):
    roll = random.randint(1, 100)
    ensure_user(user_id)
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        p = "%s" if DATABASE_URL else "?"
        now = int(time.time())
        
        if roll <= 45:
            cursor.execute(f"UPDATE users SET coin = coin + {p} WHERE user_id={p}", (500, user_id))
            result = "💰 재화 500 획득!"
        elif roll <= 70:
            cursor.execute(f"UPDATE users SET coin = coin + {p} WHERE user_id={p}", (1000, user_id))
            result = "💰 재화 1,000 획득!"
        elif roll <= 80:
            cursor.execute(f"UPDATE users SET coin = coin + {p} WHERE user_id={p}", (1500, user_id))
            result = "💰 재화 1,500 획득!"
        elif roll <= 90:
            cursor.execute(f"SELECT booster_until FROM users WHERE user_id={p}", (user_id,))
            row = cursor.fetchone()
            curr_booster = row[0] if row else 0
            new_booster = max(curr_booster, now) + 86400
            cursor.execute(f"UPDATE users SET booster_until = {p} WHERE user_id={p}", (new_booster, user_id))
            result = "💎 XP 부스터 1일 획득!"
        else:
            cursor.execute(f"UPDATE users SET premium_box = premium_box + 1 WHERE user_id={p}", (user_id,))
            result = "🎁 프리미엄 랜덤 상자 1개 획득!"
            
        conn.commit()
        cursor.close()
        conn.close()
        return result
    except Exception as e:
        print(f"❌ open_random_box 오류: {e}")
        return "상자를 여는 도중 오류가 발생했습니다."

def open_random_box_multiple(user_id: int, count: int):
    ensure_user(user_id)
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        p = "%s" if DATABASE_URL else "?"
        now = int(time.time())
        
        # Check current random boxes
        cursor.execute(f"SELECT random_box, coin, premium_box, jackpot_box, booster_until FROM users WHERE user_id={p}", (user_id,))
        row = cursor.fetchone()
        if not row or row[0] < count:  # random_box is the 1st column (index 0) in the SELECT list: random_box, coin, premium_box, jackpot_box, booster_until
            cursor.close()
            conn.close()
            return None, "보유한 랜덤 상자가 부족합니다."
            
        curr_random_box, curr_coin, curr_premium_box, curr_jackpot, curr_booster = row
        
        added_coins = 0
        added_booster_seconds = 0
        added_premium_boxes = 0
        added_jackpot_boxes = 0
        
        rewards_summary = {
            "coins": 0,
            "booster_days": 0,
            "premium_boxes": 0,
            "jackpot_boxes": 0
        }
        
        for _ in range(count):
            roll = random.randint(1, 100)
            if roll <= 45:
                added_coins += 500
                rewards_summary["coins"] += 500
            elif roll <= 70:
                added_coins += 1000
                rewards_summary["coins"] += 1000
            elif roll <= 80:
                added_coins += 1500
                rewards_summary["coins"] += 1500
            elif roll <= 90:
                added_booster_seconds += 86400
                rewards_summary["booster_days"] += 1
            else:
                added_premium_boxes += 1
                rewards_summary["premium_boxes"] += 1
                
        new_booster = max(curr_booster, now) + added_booster_seconds
        
        cursor.execute(
            f"""
            UPDATE users 
            SET random_box = random_box - {p},
                coin = coin + {p},
                premium_box = premium_box + {p},
                jackpot_box = jackpot_box + {p},
                booster_until = {p}
            WHERE user_id = {p}
            """,
            (count, added_coins, added_premium_boxes, added_jackpot_boxes, new_booster, user_id)
        )
        
        conn.commit()
        cursor.close()
        conn.close()
        return rewards_summary, None
    except Exception as e:
        print(f"❌ open_random_box_multiple 오류: {e}")
        return None, "상자를 여는 도중 오류가 발생했습니다."

def open_premium_box(user_id: int):
    roll = random.randint(1, 100)
    ensure_user(user_id)
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        p = "%s" if DATABASE_URL else "?"
        now = int(time.time())
        
        if roll <= 40:
            cursor.execute(f"UPDATE users SET coin = coin + {p} WHERE user_id={p}", (5000, user_id))
            result = "💰 재화 5,000 획득!"
        elif roll <= 65:
            cursor.execute(f"UPDATE users SET random_box = random_box + 10 WHERE user_id={p}", (user_id,))
            result = "📦 랜덤 상자 10개 획득!"
        elif roll <= 80:
            cursor.execute(f"UPDATE users SET coin = coin + {p} WHERE user_id={p}", (7500, user_id))
            result = "💰 재화 7,500 획득!"
        elif roll <= 90:
            cursor.execute(f"SELECT booster_until FROM users WHERE user_id={p}", (user_id,))
            row = cursor.fetchone()
            curr_booster = row[0] if row else 0
            new_booster = max(curr_booster, now) + 3 * 86400
            cursor.execute(f"UPDATE users SET booster_until = {p} WHERE user_id={p}", (new_booster, user_id))
            result = "💎 XP 부스터 3일 획득!"
        elif roll <= 97:
            cursor.execute(f"SELECT booster_until FROM users WHERE user_id={p}", (user_id,))
            row = cursor.fetchone()
            curr_booster = row[0] if row else 0
            new_booster = max(curr_booster, now) + 15 * 86400
            cursor.execute(f"UPDATE users SET booster_until = {p} WHERE user_id={p}", (new_booster, user_id))
            result = "💎 XP 부스터 15일 획득!"
        else:
            cursor.execute(f"UPDATE users SET jackpot_box = jackpot_box + 1 WHERE user_id={p}", (user_id,))
            result = "👑 잭팟 상자 1개 획득!"
            
        conn.commit()
        cursor.close()
        conn.close()
        return result
    except Exception as e:
        print(f"❌ open_premium_box 오류: {e}")
        return "상자를 여는 도중 오류가 발생했습니다."

def open_jackpot_box(user_id: int):
    roll = random.randint(1, 100)
    ensure_user(user_id)
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        p = "%s" if DATABASE_URL else "?"
        now = int(time.time())
        
        if roll <= 50:
            cursor.execute(f"UPDATE users SET coin = coin + {p} WHERE user_id={p}", (10000, user_id))
            result = "💰 재화 10,000 획득!"
        elif roll <= 80:
            cursor.execute(f"UPDATE users SET premium_box = premium_box + 5 WHERE user_id={p}", (user_id,))
            result = "🎁 프리미엄 랜덤 상자 5개 획득!"
        elif roll <= 95:
            cursor.execute(f"SELECT booster_until FROM users WHERE user_id={p}", (user_id,))
            row = cursor.fetchone()
            curr_booster = row[0] if row else 0
            new_booster = max(curr_booster, now) + 30 * 86400
            cursor.execute(f"UPDATE users SET booster_until = {p} WHERE user_id={p}", (new_booster, user_id))
            result = "💎 XP 부스터 30일 획득!"
        else:
            result = "🎁 기프티콘 획득! (관리자에게 문의해주세요.)"
            
        conn.commit()
        cursor.close()
        conn.close()
        return result
    except Exception as e:
        print(f"❌ open_jackpot_box 오류: {e}")
        return "상자를 여는 도중 오류가 발생했습니다."


def open_premium_box_multiple(user_id: int, count: int):
    ensure_user(user_id)
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        p = "%s" if DATABASE_URL else "?"
        now = int(time.time())
        
        # Check current premium boxes
        cursor.execute(f"SELECT premium_box, coin, random_box, jackpot_box, booster_until FROM users WHERE user_id={p}", (user_id,))
        row = cursor.fetchone()
        if not row or row[0] < count:
            cursor.close()
            conn.close()
            return None, "보유한 프리미엄 랜덤 상자가 부족합니다."
            
        curr_premium_box, curr_coin, curr_random_box, curr_jackpot, curr_booster = row
        
        added_coins = 0
        added_random_boxes = 0
        added_booster_seconds = 0
        added_jackpot_boxes = 0
        
        rewards_summary = {
            "coins": 0,
            "random_boxes": 0,
            "booster_days": 0,
            "jackpot_boxes": 0
        }
        
        for _ in range(count):
            roll = random.randint(1, 100)
            if roll <= 40:
                added_coins += 5000
                rewards_summary["coins"] += 5000
            elif roll <= 65:
                added_random_boxes += 10
                rewards_summary["random_boxes"] += 10
            elif roll <= 80:
                added_coins += 7500
                rewards_summary["coins"] += 7500
            elif roll <= 90:
                added_booster_seconds += 3 * 86400
                rewards_summary["booster_days"] += 3
            elif roll <= 97:
                added_booster_seconds += 15 * 86400
                rewards_summary["booster_days"] += 15
            else:
                added_jackpot_boxes += 1
                rewards_summary["jackpot_boxes"] += 1
                
        new_booster = max(curr_booster, now) + added_booster_seconds
        
        cursor.execute(
            f"""
            UPDATE users 
            SET premium_box = premium_box - {p},
                coin = coin + {p},
                random_box = random_box + {p},
                jackpot_box = jackpot_box + {p},
                booster_until = {p}
            WHERE user_id = {p}
            """,
            (count, added_coins, added_random_boxes, added_jackpot_boxes, new_booster, user_id)
        )
        
        conn.commit()
        cursor.close()
        conn.close()
        return rewards_summary, None
    except Exception as e:
        print(f"❌ open_premium_box_multiple 오류: {e}")
        return None, "상자를 여는 도중 오류가 발생했습니다."


def open_jackpot_box_multiple(user_id: int, count: int):
    ensure_user(user_id)
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        p = "%s" if DATABASE_URL else "?"
        now = int(time.time())
        
        # Check current jackpot boxes
        cursor.execute(f"SELECT jackpot_box, coin, premium_box, booster_until FROM users WHERE user_id={p}", (user_id,))
        row = cursor.fetchone()
        if not row or row[0] < count:
            cursor.close()
            conn.close()
            return None, "보유한 잭팟 상자가 부족합니다."
            
        curr_jackpot_box, curr_coin, curr_premium_box, curr_booster = row
        
        added_coins = 0
        added_premium_boxes = 0
        added_booster_seconds = 0
        gifticon_count = 0
        
        rewards_summary = {
            "coins": 0,
            "premium_boxes": 0,
            "booster_days": 0,
            "gifticons": 0
        }
        
        for _ in range(count):
            roll = random.randint(1, 100)
            if roll <= 50:
                added_coins += 10000
                rewards_summary["coins"] += 10000
            elif roll <= 80:
                added_premium_boxes += 5
                rewards_summary["premium_boxes"] += 5
            elif roll <= 95:
                added_booster_seconds += 30 * 86400
                rewards_summary["booster_days"] += 30
            else:
                gifticon_count += 1
                rewards_summary["gifticons"] += 1
                
        new_booster = max(curr_booster, now) + added_booster_seconds
        
        cursor.execute(
            f"""
            UPDATE users 
            SET jackpot_box = jackpot_box - {p},
                coin = coin + {p},
                premium_box = premium_box + {p},
                booster_until = {p}
            WHERE user_id = {p}
            """,
            (count, added_coins, added_premium_boxes, new_booster, user_id)
        )
        
        conn.commit()
        cursor.close()
        conn.close()
        return rewards_summary, None
    except Exception as e:
        print(f"❌ open_jackpot_box_multiple 오류: {e}")
        return None, "상자를 여는 도중 오류가 발생했습니다."


# 임베드 생성 함수들
def pass_embed(member: discord.Member):
    xp, coin, random_box, premium_box, jackpot_box, booster_until, voice_minutes = get_user(member.id)
    level, current_xp, need_xp = level_from_xp(xp)

    embed = discord.Embed(
        title="🎫 HEAVEN 시즌 패스",
        description=f"{member.mention}님의 실시간 패스 정보",
        color=0x8e44ad
    )

    embed.add_field(name="레벨", value=f"Lv.{level}", inline=True)
    embed.add_field(name="XP", value=f"{current_xp} / {need_xp}", inline=True)
    embed.add_field(name="진행도", value=progress_bar(current_xp, need_xp), inline=False)

    embed.add_field(name="💰 보유 재화", value=f"{coin:,}", inline=True)
    embed.add_field(name="🎤 누적 음성시간", value=f"{voice_minutes:,}분", inline=True)

    embed.add_field(
        name="📦 보유 상자",
        value=f"랜덤 상자: {random_box}개\n프리미엄 상자: {premium_box}개\n잭팟 상자: {jackpot_box}개",
        inline=False
    )

    now = int(time.time())
    if booster_until > now:
        booster_status = f"🔥 활성화 중 (만료: <t:{booster_until}:F> / <t:{booster_until}:R>)"
    else:
        booster_status = "❌ 비활성화"
    embed.add_field(name="💎 XP 부스터", value=booster_status, inline=False)

    embed.add_field(name="🎁 다음 보상", value=next_reward(level), inline=False)
    return embed

def quest_embed(user_id: int):
    today = get_current_date()
    ensure_user(user_id)
    
    conn = get_db_connection()
    cursor = conn.cursor()
    p = "%s" if DATABASE_URL else "?"
    
    cursor.execute(f"""
        SELECT quest_id, progress, claimed FROM user_quests
        WHERE user_id = {p} AND quest_date = {p}
    """, (user_id, today))
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    
    quest_data = {row[0]: {"progress": row[1], "claimed": row[2]} for row in rows}
    
    embed = discord.Embed(
        title="📜 오늘의 일일 퀘스트",
        description="매일 오전 6시에 초기화되는 시즌 패스 일일 미션입니다.\n미션을 달성하고 보상을 수령하세요!",
        color=0x3498db
    )
    
    for q_id, q_info in DAILY_QUESTS.items():
        db_info = quest_data.get(q_id, {"progress": 0, "claimed": 0})
        progress = min(db_info["progress"], q_info["target"])
        target = q_info["target"]
        
        status_str = ""
        if db_info["claimed"]:
            status_str = "✅ **보상 수령 완료**"
        elif progress >= target:
            status_str = "🎁 **수령 가능 (아래 일괄 수령 버튼을 누르세요)**"
        else:
            status_str = f"⚡ 진행도: `{progress}/{target}`"
            
        embed.add_field(
            name=q_info["title"],
            value=(
                f"{status_str}\n"
                f"└ 보상: ⭐ {q_info['xp_reward']} XP / 💰 {q_info['coin_reward']} 코인"
            ),
            inline=False
        )
        
    return embed

class QuestPanelView(discord.ui.View):
    def __init__(self, user_id: int):
        super().__init__(timeout=180)
        self.user_id = user_id

    @discord.ui.button(label="보상 일괄 수령", emoji="🎁", style=discord.ButtonStyle.success, custom_id="heaven_quest:claim_all")
    async def claim_all(self, interaction: discord.Interaction, button: discord.ui.Button):
        if interaction.user.id != self.user_id:
            return await interaction.response.send_message("❌ 본인의 퀘스트 보상만 수령할 수 있습니다.", ephemeral=True)
            
        await interaction.response.send_message("🎁 퀘스트 보상 가방을 여는 중... ⚙️", ephemeral=True)
        await asyncio.sleep(0.5)
        
        success, res = claim_all_quests_calc(interaction.user.id)
        if not success:
            return await interaction.edit_original_response(content=f"❌ {res}")
            
        total_xp, total_coins, claimed_quests = res
        
        await interaction.edit_original_response(content=f"💰 재화 정산 중... (+{total_coins:,} 코인) 💸")
        await asyncio.sleep(0.5)
        await interaction.edit_original_response(content=f"⭐ 경험치 획득 중... (+{total_xp:,} XP) ✨")
        await asyncio.sleep(0.5)
        
        msg = apply_claimed_quests(interaction.user.id, total_xp, total_coins, claimed_quests)
        await interaction.edit_original_response(content=msg)
        
        new_embed = quest_embed(interaction.user.id)
        await interaction.message.edit(embed=new_embed, view=self)

def shop_embed():
    embed = discord.Embed(
        title="🛒 HEAVEN 상점",
        description="버튼으로 구매할 상품을 선택하세요.",
        color=0x2ecc71
    )
    embed.add_field(name="📦 랜덤 상자", value="2,000 재화", inline=False)
    embed.add_field(name="💎 XP 부스터 1일", value="1,000 재화", inline=False)
    embed.add_field(name="💎 XP 부스터 7일", value="5,000 재화", inline=False)
    embed.add_field(name="🎁 프리미엄 랜덤 상자", value="8,000 재화", inline=False)
    return embed

def box_info_embed():
    embed = discord.Embed(
        title="📦 상자 확률표",
        color=0xf1c40f
    )
    embed.add_field(
        name="📦 랜덤 상자",
        value=(
            "45% → 💰 재화 500\n"
            "25% → 💰 재화 1,000\n"
            "10% → 💰 재화 1,500\n"
            "10% → 💎 XP 부스터 1일\n"
            "10% → 🎁 프리미엄 랜덤 상자"
        ),
        inline=False
    )
    embed.add_field(
        name="🎁 프리미엄 랜덤 상자",
        value=(
            "40% → 💰 재화 5,000\n"
            "25% → 📦 랜덤 상자 10개\n"
            "15% → 💰 재화 7,500\n"
            "10% → 💎 XP 부스터 3일\n"
            "7% → 💎 XP 부스터 15일\n"
            "3% → 👑 잭팟 상자"
        ),
        inline=False
    )
    embed.add_field(
        name="👑 잭팟 상자",
        value=(
            "50% → 💰 재화 10,000\n"
            "30% → 🎁 프리미엄 랜덤 상자 5개\n"
            "15% → 💎 XP 부스터 30일\n"
            "5% → 🎁 기프티콘"
        ),
        inline=False
    )
    return embed

def rewards_info_embed():
    embed = discord.Embed(
        title="🎫 HEAVEN 시즌 패스 전체 보상 목록",
        description="레벨 달성 시 인벤토리 및 계정에 즉시 자동 지급되는 보상들입니다.",
        color=0x9b59b6
    )
    
    reward_lines = [
        "⭐ **Lv.5** : 💰 재화 500",
        "⭐ **Lv.10** : 📦 랜덤 상자 1개",
        "⭐ **Lv.15** : 💰 재화 1,000",
        "⭐ **Lv.20** : 📦 랜덤 상자 2개",
        "⭐ **Lv.25** : 💰 재화 2,500",
        "⭐ **Lv.30** : 🎁 프리미엄 상자 1개",
        "⭐ **Lv.35** : 💰 재화 3,000",
        "⭐ **Lv.40** : 🎁 프리미엄 상자 2개",
        "⭐ **Lv.45** : 📦 랜덤 상자 5개",
        "⭐ **Lv.50** : 👑 잭팟 상자 1개"
    ]
    
    embed.description = "\n".join(reward_lines)
    return embed


# Flask Web Server to keep the bot alive
app = Flask('')

@app.route('/')
def home():
    return "Bot is alive!"

def run():
    app.run(host='0.0.0.0', port=3000)

def keep_alive():
    t = Thread(target=run)
    t.start()

# Intents configuration
intents = discord.Intents.default()
intents.message_content = True
intents.voice_states = True
intents.members = True

# Bot initialization
bot = commands.Bot(command_prefix=commands.when_mentioned, intents=intents)

class PassRankingView(discord.ui.View):
    def __init__(self, full_rows):
        super().__init__(timeout=180)  # 3분 제한
        self.full_rows = full_rows

    @discord.ui.button(label="더보기", style=discord.ButtonStyle.primary, custom_id="show_more_pass_ranking")
    async def show_more(self, interaction: discord.Interaction, button: discord.ui.Button):
        embed = discord.Embed(
            title="🏆 HEAVEN 시즌 패스 전체 랭킹",
            color=0x8e44ad
        )
        
        desc_lines = []
        # 디스코드 글자 수 제한(4096자)을 방지하기 위해 상위 50명까지 표시
        limit_rows = self.full_rows[:50]
        for idx, (user_id, xp) in enumerate(limit_rows, 1):
            level, _, _ = level_from_xp(xp)
            desc_lines.append(f"{idx}등: <@{user_id}> - Lv.{level} ({xp:,} XP)")
            
        if len(self.full_rows) > 50:
            desc_lines.append("\n*상위 50명까지 표시됩니다.*")
            
        embed.description = "\n".join(desc_lines)
        
        # 버튼 제거
        self.clear_items()
        await interaction.response.edit_message(embed=embed, view=self)

class VoiceUsageView(discord.ui.View):
    def __init__(self, full_rows):
        super().__init__(timeout=86400)  # 24시간
        self.full_rows = full_rows

    @discord.ui.button(label="나머지 보기", style=discord.ButtonStyle.primary, custom_id="show_more_voice_usage")
    async def show_more(self, interaction: discord.Interaction, button: discord.ui.Button):
        embed = discord.Embed(
            title="📅 음성 채널 이용 시간 전체 랭킹",
            color=discord.Color.blue()
        )
        
        desc_lines = []
        limit_rows = self.full_rows[:50]
        for idx, (user_id, seconds) in enumerate(limit_rows, 1):
            desc_lines.append(f"{idx}등: <@{user_id}> - {format_time(seconds)}")
            
        if len(self.full_rows) > 50:
            desc_lines.append("\n*상위 50명까지 표시됩니다.*")
            
        embed.description = "\n".join(desc_lines)
        
        # 버튼 제거
        self.clear_items()
        await interaction.response.edit_message(embed=embed, view=self)

class DateSelect(discord.ui.Select):
    def __init__(self, dates, placeholder="조회할 날짜를 선택하세요..."):
        options = [
            discord.SelectOption(label=f"{d}", value=d)
            for d in dates
        ]
        super().__init__(placeholder=placeholder, min_values=1, max_values=1, options=options)

    async def callback(self, interaction: discord.Interaction):
        selected_date = self.values[0]
        
        conn = get_db_connection()
        cursor = conn.cursor()
        p = "%s" if DATABASE_URL else "?"
        cursor.execute(f"SELECT user_id, seconds FROM voice_usage WHERE use_date = {p} ORDER BY seconds DESC", (selected_date,))
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        
        if not rows:
            embed = discord.Embed(
                title=f"📅 {selected_date} 음성 채널 이용 시간 랭킹",
                description=f"{selected_date}에 음성 채널을 이용한 유저가 없습니다.",
                color=discord.Color.orange()
            )
            await interaction.response.edit_message(embed=embed, view=None)
            return
            
        embed = discord.Embed(
            title=f"📅 {selected_date} 음성 채널 이용 시간 랭킹",
            color=discord.Color.green()
        )
        
        top_5 = rows[:5]
        desc_lines = []
        for idx, (user_id, seconds) in enumerate(top_5, 1):
            desc_lines.append(f"{idx}등: <@{user_id}> - {format_time(seconds)}")
            
        if len(rows) > 5:
            desc_lines.append("\n*6등 이하의 기록은 아래 버튼을 눌러 확인하세요.*")
            embed.description = "\n".join(desc_lines)
            view = VoiceUsageView(rows)
            await interaction.response.edit_message(embed=embed, view=view)
        else:
            embed.description = "\n".join(desc_lines)
            await interaction.response.edit_message(embed=embed, view=None)

class DateSelectView(discord.ui.View):
    def __init__(self, dates):
        super().__init__(timeout=180)  # 3분 제한
        # 디스코드 선택 메뉴는 최대 25개 옵션까지 지원하므로, 날짜 데이터를 20개 단위로 쪼개어 드롭다운을 생성합니다.
        # 최대 5개의 드롭다운(총 100일 분량)까지 한 메시지에 등록 가능합니다.
        chunk_size = 20
        visible_dates = dates[:100]
        
        for i in range(0, len(visible_dates), chunk_size):
            chunk = visible_dates[i:i+chunk_size]
            start_num = i + 1
            end_num = i + len(chunk)
            if i == 0:
                placeholder = f"최근 날짜 선택 (1~{end_num}일)..."
            else:
                placeholder = f"이전 날짜 선택 ({start_num}~{end_num}일)..."
            self.add_item(DateSelect(chunk, placeholder=placeholder))

class VoiceUsagePanel(discord.ui.View):
    def __init__(self):
        super().__init__(timeout=None)  # 영구적인 뷰

    @discord.ui.button(label="오늘의 랭킹", style=discord.ButtonStyle.success, custom_id="check_voice_today_btn")
    async def check_today(self, interaction: discord.Interaction, button: discord.ui.Button):
        today_str = get_current_date()
        rows = get_realtime_today_stats()
        
        if not rows:
            embed = discord.Embed(
                title=f"📅 오늘의 음성 채널 이용 시간 랭킹 ({today_str})",
                description="오늘 음성 채널을 이용한 유저가 없습니다.",
                color=discord.Color.orange()
            )
            await interaction.response.send_message(embed=embed, ephemeral=True)
            return
            
        embed = discord.Embed(
            title=f"📅 오늘의 음성 채널 이용 시간 랭킹 ({today_str})",
            color=discord.Color.green()
        )
        
        top_5 = rows[:5]
        desc_lines = []
        for idx, (user_id, seconds) in enumerate(top_5, 1):
            desc_lines.append(f"{idx}등: <@{user_id}> - {format_time(seconds)}")
            
        if len(rows) > 5:
            desc_lines.append("\n*6등 이하의 기록은 아래 버튼을 눌러 확인하세요.*")
            embed.description = "\n".join(desc_lines)
            view = VoiceUsageView(rows)
            await interaction.response.send_message(embed=embed, view=view, ephemeral=True)
        else:
            embed.description = "\n".join(desc_lines)
            await interaction.response.send_message(embed=embed, ephemeral=True)

    @discord.ui.button(label="누적 전체 랭킹", style=discord.ButtonStyle.primary, custom_id="check_voice_cumulative_btn")
    async def check_cumulative(self, interaction: discord.Interaction, button: discord.ui.Button):
        rows = get_realtime_cumulative_stats()
        
        if not rows:
            embed = discord.Embed(
                title="🏆 누적 음성 채널 이용 시간 랭킹",
                description="누적된 음성 채널 이용 기록이 없습니다.",
                color=discord.Color.orange()
            )
            await interaction.response.send_message(embed=embed, ephemeral=True)
            return
            
        embed = discord.Embed(
            title="🏆 누적 음성 채널 이용 시간 랭킹",
            color=discord.Color.gold()
        )
        
        top_5 = rows[:5]
        desc_lines = []
        for idx, (user_id, seconds) in enumerate(top_5, 1):
            desc_lines.append(f"{idx}등: <@{user_id}> - {format_time(seconds)}")
            
        if len(rows) > 5:
            desc_lines.append("\n*6등 이하의 기록은 아래 버튼을 눌러 확인하세요.*")
            embed.description = "\n".join(desc_lines)
            view = VoiceUsageView(rows)
            await interaction.response.send_message(embed=embed, view=view, ephemeral=True)
        else:
            embed.description = "\n".join(desc_lines)
            await interaction.response.send_message(embed=embed, ephemeral=True)

    @discord.ui.button(label="날짜별 랭킹 조회", style=discord.ButtonStyle.secondary, custom_id="check_voice_select_date_btn")
    async def check_by_date(self, interaction: discord.Interaction, button: discord.ui.Button):
        dates = get_recent_dates()
        
        if not dates:
            embed = discord.Embed(
                title="📅 날짜별 랭킹 조회",
                description="조회 가능한 음성 채널 이용 기록이 데이터베이스에 존재하지 않습니다.",
                color=discord.Color.orange()
            )
            await interaction.response.send_message(embed=embed, ephemeral=True)
            return
            
        embed = discord.Embed(
            title="📅 날짜별 랭킹 조회",
            description="아래 선택 메뉴에서 조회할 날짜를 선택해주세요.",
            color=discord.Color.blue()
        )
        view = DateSelectView(dates)
        await interaction.response.send_message(embed=embed, view=view, ephemeral=True)

    @discord.ui.button(label="월간 기록 & 미접속자", style=discord.ButtonStyle.success, custom_id="check_voice_monthly_btn")
    async def check_monthly(self, interaction: discord.Interaction, button: discord.ui.Button):
        now = get_kst_now()
        embed = build_monthly_dashboard_embed(interaction.guild, now.year, now.month)
        view = MonthlyDashboardView(interaction.guild, now.year, now.month)
        await interaction.response.send_message(embed=embed, view=view, ephemeral=True)


# ==========================================
# [월간 활동 및 미접속자 프리미엄 대시보드 UI]
# ==========================================

def build_monthly_dashboard_embed(guild: discord.Guild, year: int, month: int):
    """고급스러운 월간 활동 및 미접속자 통합 대시보드 Embed 생성"""
    stats = get_monthly_voice_stats(year, month)
    inactive_list = get_inactive_members(guild, days_threshold=14) if guild else []
    
    total_members = len([m for m in guild.members if not m.bot]) if guild else 0
    active_users_count = stats["total_users"]
    total_seconds = stats["total_seconds"]
    
    server_title = guild.name if guild else "HEAVEN"
    embed = discord.Embed(
        title=f"👑 【 {server_title} 】 {year}년 {month:02d}월 종합 활동 리포트",
        description=(
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"📅 **집계 대상 기간**: `{stats['start_date']} ~ {stats['end_date']}`\n"
            f"💡 *매월 1일 00:00부터 말일 23:59까지의 기록이 자동으로 집계·보존됩니다.*"
        ),
        color=0x5865F2
    )
    
    # 1. 월간 활동 통계 요약
    avg_seconds = (total_seconds // active_users_count) if active_users_count > 0 else 0
    active_rate = (active_users_count / total_members * 100) if total_members > 0 else 0
    embed.add_field(
        name="📊 월간 음성 활동 요약",
        value=(
            f"• **음성 참여 멤버**: `{active_users_count:,}명` / `{total_members:,}명` ({active_rate:.1f}%)\n"
            f"• **총 이용 시간**: `{format_time(total_seconds)}`\n"
            f"• **1인 평균 시간**: `{format_time(avg_seconds)}`"
        ),
        inline=False
    )
    
    # 2. 월간 랭킹 TOP 5
    top_5 = stats["rankings"][:5]
    if top_5:
        medals = ["🥇", "🥈", "🥉", "🎖️", "🎖️"]
        top_lines = []
        for idx, (uid, secs) in enumerate(top_5, 1):
            medal = medals[idx-1] if idx <= len(medals) else f"`{idx}등`"
            top_lines.append(f"{medal} **{idx}등** <@{uid}> — `{format_time(secs)}`")
        embed.add_field(
            name=f"🏆 {month}월 음성 랭킹 TOP 5",
            value="\n".join(top_lines),
            inline=False
        )
    else:
        embed.add_field(
            name=f"🏆 {month}월 음성 랭킹 TOP 5",
            value="*해당 월에는 아직 기록된 음성 활동이 없습니다.*",
            inline=False
        )
        
    # 3. 2주 이상 미접속 멤버 요약
    inactive_count = len(inactive_list)
    inactive_rate = (inactive_count / total_members * 100) if total_members > 0 else 0
    
    inactive_preview = []
    for item in inactive_list[:4]:
        status_type = "최근 음성" if item["has_record"] else "가입일"
        inactive_preview.append(
            f"• **{item['display_name']}** (<@{item['user_id']}>) ➔ ⚠️ **{item['inactive_days']}일 미접속** `({status_type}: {item['last_date_str']})`"
        )
    
    preview_str = "\n".join(inactive_preview) if inactive_preview else "• *현재 2주(14일) 이상 미접속 멤버가 없습니다.*"
    if inactive_count > 4:
        preview_str += f"\n*... 외 {inactive_count - 4}명 (아래 [2주+ 미접속자 명단] 버튼으로 확인)*"
        
    embed.add_field(
        name=f"🚨 2주(14일) 이상 미접속 멤버 ({inactive_count}명 / {inactive_rate:.1f}%)",
        value=preview_str,
        inline=False
    )
    
    embed.set_footer(text=f"기준 시각: {get_kst_now().strftime('%Y-%m-%d %H:%M:%S')} (KST) • 매월 1일~말일 자동 갱신")
    return embed


class MonthDropdown(discord.ui.Select):
    def __init__(self, guild, current_year, current_month):
        self.guild = guild
        months = get_available_months()
        options = []
        cur_val = f"{current_year:04d}-{current_month:02d}"
        for ym in months:
            y, m = ym.split("-")
            label = f"{y}년 {int(m):02d}월 종합 기록"
            options.append(discord.SelectOption(
                label=label,
                value=ym,
                default=(ym == cur_val),
                description=f"{y}년 {int(m):02d}월 1일 ~ 말일 활동 내역"
            ))
        super().__init__(
            placeholder="조회할 연월을 선택하세요...",
            min_values=1,
            max_values=1,
            options=options
        )

    async def callback(self, interaction: discord.Interaction):
        selected_ym = self.values[0]
        y, m = map(int, selected_ym.split("-"))
        embed = build_monthly_dashboard_embed(self.guild, y, m)
        view = MonthlyDashboardView(self.guild, y, m)
        await interaction.response.edit_message(embed=embed, view=view)


class MonthlyDashboardView(discord.ui.View):
    def __init__(self, guild: discord.Guild, year: int, month: int):
        super().__init__(timeout=300)
        self.guild = guild
        self.year = year
        self.month = month
        
        self.add_item(MonthDropdown(guild, year, month))

    @discord.ui.button(label="월간 랭킹 전체보기", style=discord.ButtonStyle.primary, emoji="🏆")
    async def view_full_ranking(self, interaction: discord.Interaction, button: discord.ui.Button):
        stats = get_monthly_voice_stats(self.year, self.month)
        if not stats["rankings"]:
            await interaction.response.send_message(f"⚠️ {self.year}년 {self.month}월에는 기록된 활동 랭킹이 없습니다.", ephemeral=True)
            return
        ranking_view = MonthlyRankingView(self.guild, self.year, self.month, stats["rankings"])
        embed = ranking_view.create_embed()
        await interaction.response.edit_message(embed=embed, view=ranking_view)

    @discord.ui.button(label="2주+ 미접속자 명단", style=discord.ButtonStyle.danger, emoji="🚨")
    async def view_inactive_members(self, interaction: discord.Interaction, button: discord.ui.Button):
        inactive_list = get_inactive_members(self.guild, days_threshold=14) if self.guild else []
        if not inactive_list:
            await interaction.response.send_message("✨ 2주(14일) 이상 미접속 멤버가 없습니다! 모든 멤버가 활동 중입니다.", ephemeral=True)
            return
        inactive_view = InactiveMembersView(self.guild, self.year, self.month, inactive_list)
        embed = inactive_view.create_embed()
        await interaction.response.edit_message(embed=embed, view=inactive_view)

    @discord.ui.button(label="새로고침", style=discord.ButtonStyle.secondary, emoji="🔄")
    async def refresh_dashboard(self, interaction: discord.Interaction, button: discord.ui.Button):
        embed = build_monthly_dashboard_embed(self.guild, self.year, self.month)
        view = MonthlyDashboardView(self.guild, self.year, self.month)
        await interaction.response.edit_message(embed=embed, view=view)


class MonthlyRankingView(discord.ui.View):
    def __init__(self, guild, year, month, rankings):
        super().__init__(timeout=300)
        self.guild = guild
        self.year = year
        self.month = month
        self.rankings = rankings
        self.current_page = 0
        self.per_page = 15
        self.total_pages = (len(rankings) - 1) // self.per_page + 1
        self.update_buttons()

    def update_buttons(self):
        self.clear_items()
        
        prev_btn = discord.ui.Button(label="◀ 이전", style=discord.ButtonStyle.secondary, disabled=(self.current_page == 0))
        prev_btn.callback = self.prev_page
        self.add_item(prev_btn)

        page_btn = discord.ui.Button(label=f"{self.current_page + 1} / {self.total_pages}", style=discord.ButtonStyle.secondary, disabled=True)
        self.add_item(page_btn)

        next_btn = discord.ui.Button(label="다음 ▶", style=discord.ButtonStyle.secondary, disabled=(self.current_page >= self.total_pages - 1))
        next_btn.callback = self.next_page
        self.add_item(next_btn)

        back_btn = discord.ui.Button(label="🔙 대시보드로 돌아가기", style=discord.ButtonStyle.primary)
        back_btn.callback = self.go_back
        self.add_item(back_btn)

    def create_embed(self):
        embed = discord.Embed(
            title=f"🏆 {self.year}년 {self.month:02d}월 음성 이용 시간 전체 랭킹",
            description=f"총 활동 유저: **{len(self.rankings):,}명** (페이지 {self.current_page + 1}/{self.total_pages})",
            color=0xFEE75C
        )
        start_idx = self.current_page * self.per_page
        page_items = self.rankings[start_idx : start_idx + self.per_page]
        
        lines = []
        for idx, (uid, secs) in enumerate(page_items, start_idx + 1):
            lines.append(f"**{idx}등** <@{uid}> — `{format_time(secs)}`")
            
        embed.description = f"총 활동 유저: **{len(self.rankings):,}명**\n\n" + "\n".join(lines)
        embed.set_footer(text=f"{self.year}년 {self.month:02d}월 종합 집계")
        return embed

    async def prev_page(self, interaction: discord.Interaction):
        if self.current_page > 0:
            self.current_page -= 1
            self.update_buttons()
            await interaction.response.edit_message(embed=self.create_embed(), view=self)

    async def next_page(self, interaction: discord.Interaction):
        if self.current_page < self.total_pages - 1:
            self.current_page += 1
            self.update_buttons()
            await interaction.response.edit_message(embed=self.create_embed(), view=self)

    async def go_back(self, interaction: discord.Interaction):
        embed = build_monthly_dashboard_embed(self.guild, self.year, self.month)
        view = MonthlyDashboardView(self.guild, self.year, self.month)
        await interaction.response.edit_message(embed=embed, view=view)


class InactiveMembersView(discord.ui.View):
    def __init__(self, guild, year, month, inactive_list):
        super().__init__(timeout=300)
        self.guild = guild
        self.year = year
        self.month = month
        self.inactive_list = inactive_list
        self.current_page = 0
        self.per_page = 15
        self.total_pages = (len(inactive_list) - 1) // self.per_page + 1 if inactive_list else 1
        self.update_buttons()

    def update_buttons(self):
        self.clear_items()
        
        prev_btn = discord.ui.Button(label="◀ 이전", style=discord.ButtonStyle.secondary, disabled=(self.current_page == 0))
        prev_btn.callback = self.prev_page
        self.add_item(prev_btn)

        page_btn = discord.ui.Button(label=f"{self.current_page + 1} / {self.total_pages}", style=discord.ButtonStyle.secondary, disabled=True)
        self.add_item(page_btn)

        next_btn = discord.ui.Button(label="다음 ▶", style=discord.ButtonStyle.secondary, disabled=(self.current_page >= self.total_pages - 1))
        next_btn.callback = self.next_page
        self.add_item(next_btn)

        back_btn = discord.ui.Button(label="🔙 대시보드로 돌아가기", style=discord.ButtonStyle.primary)
        back_btn.callback = self.go_back
        self.add_item(back_btn)

    def create_embed(self):
        embed = discord.Embed(
            title="🚨 2주(14일) 이상 미접속 멤버 명단",
            description=f"최근 14일 이상 음성 채널에 들어오지 않은 멤버 목록입니다.\n총 대상자: **{len(self.inactive_list):,}명** (페이지 {self.current_page + 1}/{self.total_pages})\n\n",
            color=0xED4245
        )
        start_idx = self.current_page * self.per_page
        page_items = self.inactive_list[start_idx : start_idx + self.per_page]
        
        lines = []
        for idx, item in enumerate(page_items, start_idx + 1):
            status_type = "최근 음성" if item["has_record"] else "가입일"
            lines.append(
                f"`{idx:02d}.` **{item['display_name']}** (<@{item['user_id']}>) ➔ ⚠️ **{item['inactive_days']}일 미접속** `({status_type}: {item['last_date_str']})`"
            )
            
        embed.description = f"최근 14일 이상 음성 채널 미접속 멤버 목록입니다.\n총 인원: **{len(self.inactive_list):,}명** (오래 안 들어온 순 정렬)\n\n" + "\n".join(lines)
        embed.set_footer(text="닉네임 옆에 미접속 일수 및 최근 활동일(또는 가입일)이 표기됩니다.")
        return embed

    async def prev_page(self, interaction: discord.Interaction):
        if self.current_page > 0:
            self.current_page -= 1
            self.update_buttons()
            await interaction.response.edit_message(embed=self.create_embed(), view=self)

    async def next_page(self, interaction: discord.Interaction):
        if self.current_page < self.total_pages - 1:
            self.current_page += 1
            self.update_buttons()
            await interaction.response.edit_message(embed=self.create_embed(), view=self)

    async def go_back(self, interaction: discord.Interaction):
        embed = build_monthly_dashboard_embed(self.guild, self.year, self.month)
        view = MonthlyDashboardView(self.guild, self.year, self.month)
        await interaction.response.edit_message(embed=embed, view=view)


# =========================
# 상자 일괄 개봉을 위한 Select Menu 및 View
# =========================
class BoxOpenSelect(discord.ui.Select):
    def __init__(self, box_count: int, box_type: str):
        self.box_type = box_type  # "random", "premium", "jackpot"
        
        # Determine emoji and labels based on box type
        if box_type == "random":
            emoji = "📦"
            label_text = "랜덤 상자"
        elif box_type == "premium":
            emoji = "🎁"
            label_text = "프리미엄 상자"
        else:
            emoji = "👑"
            label_text = "잭팟 상자"
            
        options = []
        if box_count <= 25:
            for i in range(1, box_count + 1):
                options.append(discord.SelectOption(
                    label=f"{i}개",
                    value=str(i),
                    emoji=emoji,
                    description=f"{label_text} {i}개를 엽니다."
                ))
        else:
            standard_options = [1, 2, 3, 4, 5, 10, 15, 20, 25, 30, 40, 50, 60, 70, 80, 90, 100, 150, 200, 250, 300, 400, 500]
            options_to_add = [val for val in standard_options if val < box_count]
            
            for val in options_to_add:
                options.append(discord.SelectOption(
                    label=f"{val}개",
                    value=str(val),
                    emoji=emoji,
                    description=f"{label_text} {val}개를 엽니다."
                ))
                
            if len(options) >= 25:
                options = options[:24]
                
            options.append(discord.SelectOption(
                label=f"모두 열기 ({box_count}개)",
                value=str(box_count),
                emoji="🔥",
                description=f"보유 중인 {box_count}개의 상자를 모두 엽니다."
            ))
            
        super().__init__(
            placeholder="열고 싶은 상자의 개수를 선택하세요...",
            min_values=1,
            max_values=1,
            options=options
        )

    async def callback(self, interaction: discord.Interaction):
        self.disabled = True
        
        emoji_open = "📦" if self.box_type == "random" else ("🎁" if self.box_type == "premium" else "👑")
        await interaction.response.edit_message(content=f"{emoji_open} 상자를 개봉하는 중입니다... 잠시만 기다려주세요.", view=self.view)
        
        count = int(self.values[0])
        
        if self.box_type == "random":
            rewards, error_msg = open_random_box_multiple(interaction.user.id, count)
            box_name = "랜덤 상자"
            embed_color = 0x9b59b6
        elif self.box_type == "premium":
            rewards, error_msg = open_premium_box_multiple(interaction.user.id, count)
            box_name = "프리미엄 랜덤 상자"
            embed_color = 0xe74c3c
        else:
            rewards, error_msg = open_jackpot_box_multiple(interaction.user.id, count)
            box_name = "잭팟 상자"
            embed_color = 0xf1c40f
            
        if error_msg:
            return await interaction.followup.send(f"❌ {error_msg}", ephemeral=True)
            
        update_quest_progress(interaction.user.id, "open_box", count)
        
        await interaction.edit_original_response(content=f"{emoji_open} 흔들흔들... 상자들이 일제히 빛나기 시작합니다! 💫", view=None)
        await asyncio.sleep(0.5)
        await interaction.edit_original_response(content="✨ 눈부신 빛의 기둥과 함께 모든 보상이 쏟아져 나옵니다! ✨")
        await asyncio.sleep(0.5)
        
        desc_parts = [f"축하합니다! 상자 {count}개에서 다음 보상들을 획득했습니다:\n"]
        
        if self.box_type == "random":
            if rewards["coins"] > 0:
                desc_parts.append(f"* 💰 **재화 {rewards['coins']:,} 코인**")
            if rewards["booster_days"] > 0:
                desc_parts.append(f"* 💎 **XP 부스터 {rewards['booster_days']}일권**")
            if rewards["premium_boxes"] > 0:
                desc_parts.append(f"* 🎁 **프리미엄 랜덤 상자 {rewards['premium_boxes']}개**")
            if rewards["jackpot_boxes"] > 0:
                desc_parts.append(f"* 👑 **잭팟 상자 {rewards['jackpot_boxes']}개**")
        elif self.box_type == "premium":
            if rewards["coins"] > 0:
                desc_parts.append(f"* 💰 **재화 {rewards['coins']:,} 코인**")
            if rewards["random_boxes"] > 0:
                desc_parts.append(f"* 📦 **랜덤 상자 {rewards['random_boxes']}개**")
            if rewards["booster_days"] > 0:
                desc_parts.append(f"* 💎 **XP 부스터 {rewards['booster_days']}일권**")
            if rewards["jackpot_boxes"] > 0:
                desc_parts.append(f"* 👑 **잭팟 상자 {rewards['jackpot_boxes']}개**")
        else: # jackpot
            if rewards["coins"] > 0:
                desc_parts.append(f"* 💰 **재화 {rewards['coins']:,} 코인**")
            if rewards["premium_boxes"] > 0:
                desc_parts.append(f"* 🎁 **프리미엄 랜덤 상자 {rewards['premium_boxes']}개**")
            if rewards["booster_days"] > 0:
                desc_parts.append(f"* 💎 **XP 부스터 {rewards['booster_days']}일권**")
            if rewards["gifticons"] > 0:
                desc_parts.append(f"* 🎁 **기프티콘 {rewards['gifticons']}개 (관리자에게 문의해주세요.)**")
                # Send gifticon notification to admin channel
                channel_id = 1518304536136253674
                try:
                    channel = bot.get_channel(channel_id) or await bot.fetch_channel(channel_id)
                    if channel:
                        await channel.send(f"🎉 **[기프티콘 당첨]** {interaction.user.mention}님이 잭팟 상자 일괄 개봉({count}개) 중 **기프티콘 {rewards['gifticons']}개**에 당첨되었습니다! (관리자분들은 확인 후 기프티콘을 지급해 주세요.)")
                except Exception as e:
                    print(f"❌ 기프티콘 당첨 알림 전송 실패: {e}")
                    
        if len(desc_parts) == 1:
            desc_parts.append("* 꽝 (보상이 없습니다)")
            
        embed = discord.Embed(
            title=f"{emoji_open} {box_name} 일괄 개봉 완료",
            description="\n".join(desc_parts),
            color=embed_color
        )
        await interaction.edit_original_response(content=None, embed=embed)

class BoxOpenSelectView(discord.ui.View):
    def __init__(self, box_count: int, box_type: str):
        super().__init__(timeout=60)
        self.add_item(BoxOpenSelect(box_count, box_type))


# =========================
# 버튼 View (시즌 패스 및 상점)
# =========================
class PassPanelView(discord.ui.View):
    def __init__(self):
        super().__init__(timeout=None)

    @discord.ui.button(label="내 패스 보기", emoji="🎫", style=discord.ButtonStyle.primary, custom_id="heaven_pass:my_pass", row=0)
    async def my_pass(self, interaction: discord.Interaction, button: discord.ui.Button):
        await interaction.response.send_message(
            embed=pass_embed(interaction.user),
            ephemeral=True
        )

    @discord.ui.button(label="상점 보기", emoji="🛒", style=discord.ButtonStyle.success, custom_id="heaven_pass:shop", row=0)
    async def shop(self, interaction: discord.Interaction, button: discord.ui.Button):
        await interaction.response.send_message(
            embed=shop_embed(),
            view=ShopPanelView(),
            ephemeral=True
        )

    @discord.ui.button(label="랭킹 보기", emoji="🏆", style=discord.ButtonStyle.primary, custom_id="heaven_pass:ranking", row=0)
    async def ranking(self, interaction: discord.Interaction, button: discord.ui.Button):
        rows = get_season_pass_rankings()
        if not rows:
            embed = discord.Embed(
                title="🏆 HEAVEN 시즌 패스 랭킹",
                description="시즌 패스 랭킹 기록이 없습니다.",
                color=0x8e44ad
            )
            await interaction.response.send_message(embed=embed, ephemeral=True)
            return

        embed = discord.Embed(
            title="🏆 HEAVEN 시즌 패스 랭킹 (Top 5)",
            color=0x8e44ad
        )
        
        top_5 = rows[:5]
        desc_lines = []
        for idx, (user_id, xp) in enumerate(top_5, 1):
            level, _, _ = level_from_xp(xp)
            desc_lines.append(f"{idx}등: <@{user_id}> - Lv.{level} ({xp:,} XP)")
            
        if len(rows) > 5:
            desc_lines.append("\n*6등 이하의 기록은 아래 버튼을 눌러 확인하세요.*")
            embed.description = "\n".join(desc_lines)
            view = PassRankingView(rows)
            await interaction.response.send_message(embed=embed, view=view, ephemeral=True)
        else:
            embed.description = "\n".join(desc_lines)
            await interaction.response.send_message(embed=embed, ephemeral=True)

    @discord.ui.button(label="일일 퀘스트", emoji="📋", style=discord.ButtonStyle.success, custom_id="heaven_pass:quests", row=0)
    async def quests(self, interaction: discord.Interaction, button: discord.ui.Button):
        await interaction.response.send_message(
            embed=quest_embed(interaction.user.id),
            view=QuestPanelView(interaction.user.id),
            ephemeral=True
        )

    @discord.ui.button(label="랜덤 상자 열기", emoji="📦", style=discord.ButtonStyle.secondary, custom_id="heaven_pass:open_random", row=1)
    async def open_random(self, interaction: discord.Interaction, button: discord.ui.Button):
        row = get_user(interaction.user.id)
        random_box = row[2] if row else 0
        
        if random_box <= 0:
            return await interaction.response.send_message("❌ 보유한 랜덤 상자가 없습니다.", ephemeral=True)
            
        view = BoxOpenSelectView(random_box, "random")
        await interaction.response.send_message(
            f"📦 **랜덤 상자 개봉**\n개봉할 상자 개수를 선택해주세요. (보유 중: `{random_box}`개)",
            view=view,
            ephemeral=True
        )

    @discord.ui.button(label="프리미엄 상자 열기", emoji="🎁", style=discord.ButtonStyle.danger, custom_id="heaven_pass:open_premium", row=1)
    async def open_premium(self, interaction: discord.Interaction, button: discord.ui.Button):
        row = get_user(interaction.user.id)
        premium_box = row[3] if row else 0
        
        if premium_box <= 0:
            return await interaction.response.send_message("❌ 보유한 프리미엄 랜덤 상자가 없습니다.", ephemeral=True)
            
        view = BoxOpenSelectView(premium_box, "premium")
        await interaction.response.send_message(
            f"🎁 **프리미엄 상자 개봉**\n개봉할 상자 개수를 선택해주세요. (보유 중: `{premium_box}`개)",
            view=view,
            ephemeral=True
        )

    @discord.ui.button(label="잭팟 상자 열기", emoji="👑", style=discord.ButtonStyle.primary, custom_id="heaven_pass:open_jackpot", row=1)
    async def open_jackpot(self, interaction: discord.Interaction, button: discord.ui.Button):
        row = get_user(interaction.user.id)
        jackpot_box = row[4] if row else 0
        
        if jackpot_box <= 0:
            return await interaction.response.send_message("❌ 보유한 잭팟 상자가 없습니다.", ephemeral=True)
            
        view = BoxOpenSelectView(jackpot_box, "jackpot")
        await interaction.response.send_message(
            f"👑 **잭팟 상자 개봉**\n개봉할 상자 개수를 선택해주세요. (보유 중: `{jackpot_box}`개)",
            view=view,
            ephemeral=True
        )


# =========================
# 상점 일괄 구매를 위한 Select Menu 및 View
# =========================
class ShopBuySelect(discord.ui.Select):
    def __init__(self, item_type: str, base_cost: int, item_name: str):
        self.item_type = item_type
        self.base_cost = base_cost
        self.item_name = item_name
        
        options = []
        for i in range(1, 11):
            total_cost = base_cost * i
            options.append(discord.SelectOption(
                label=f"{i}개",
                value=str(i),
                description=f"구매 비용: {total_cost:,} 재화"
            ))
            
        super().__init__(
            placeholder=f"구매할 {item_name}의 개수를 선택하세요...",
            min_values=1,
            max_values=1,
            options=options
        )

    async def callback(self, interaction: discord.Interaction):
        self.disabled = True
        await interaction.response.edit_message(content="🪙 구매를 처리하는 중입니다...", view=self.view)
        
        count = int(self.values[0])
        success, msg = buy_shop_item(interaction.user.id, self.item_type, self.base_cost, count)
        
        if success:
            await interaction.edit_original_response(content=f"✅ {msg}", view=None)
        else:
            await interaction.edit_original_response(content=f"❌ {msg}", view=None)

class ShopBuySelectView(discord.ui.View):
    def __init__(self, item_type: str, base_cost: int, item_name: str):
        super().__init__(timeout=60)
        self.add_item(ShopBuySelect(item_type, base_cost, item_name))


class ShopPanelView(discord.ui.View):
    def __init__(self):
        super().__init__(timeout=None)

    @discord.ui.button(label="📦 랜덤 상자 구매", style=discord.ButtonStyle.success, custom_id="heaven_shop:buy_random")
    async def buy_random(self, interaction: discord.Interaction, button: discord.ui.Button):
        view = ShopBuySelectView("random_box", 2000, "📦 랜덤 상자")
        await interaction.response.send_message("구매할 개수를 선택해주세요.", view=view, ephemeral=True)

    @discord.ui.button(label="💎 부스터 1일 구매", style=discord.ButtonStyle.success, custom_id="heaven_shop:buy_booster_1d")
    async def buy_booster_1d(self, interaction: discord.Interaction, button: discord.ui.Button):
        view = ShopBuySelectView("booster_1d", 1000, "💎 부스터 1일")
        await interaction.response.send_message("구매할 개수를 선택해주세요.", view=view, ephemeral=True)

    @discord.ui.button(label="💎 부스터 7일 구매", style=discord.ButtonStyle.success, custom_id="heaven_shop:buy_booster_7d")
    async def buy_booster_7d(self, interaction: discord.Interaction, button: discord.ui.Button):
        view = ShopBuySelectView("booster_7d", 5000, "💎 부스터 7일")
        await interaction.response.send_message("구매할 개수를 선택해주세요.", view=view, ephemeral=True)

    @discord.ui.button(label="🎁 프리미엄 상자 구매", style=discord.ButtonStyle.danger, custom_id="heaven_shop:buy_premium")
    async def buy_premium(self, interaction: discord.Interaction, button: discord.ui.Button):
        view = ShopBuySelectView("premium_box", 8000, "🎁 프리미엄 상자")
        await interaction.response.send_message("구매할 개수를 선택해주세요.", view=view, ephemeral=True)


class StatusNicknameView(discord.ui.View):
    def __init__(self):
        super().__init__(timeout=None)

    @discord.ui.button(label="관전", style=discord.ButtonStyle.primary, custom_id="status_nick_spectate")
    async def spectate_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        try:
            member = interaction.user
            if not isinstance(member, discord.Member):
                await interaction.response.send_message("❌ 이 버튼은 서버 내에서만 사용할 수 있습니다.", ephemeral=True)
                return

            current_name = member.nick or member.name
            clean_name = current_name
            # 기존 접두사 제거
            prefixes = ["[관전] ", "[대기] ", "[관전]", "[대기]"]
            for prefix in prefixes:
                if clean_name.startswith(prefix):
                    clean_name = clean_name[len(prefix):]
                    break
            clean_name = clean_name.strip()
            new_nick = f"[관전] {clean_name}"[:32]

            try:
                await member.edit(nick=new_nick)
                await interaction.response.send_message(f"✅ 닉네임이 **{new_nick}**(으)로 변경되었습니다.", ephemeral=True)
            except discord.Forbidden:
                await interaction.response.send_message(
                    "❌ 닉네임을 변경할 권한이 없습니다.\n"
                    "(봇의 역할 서열이 본인보다 낮거나, 본인이 서버 소유자이거나, 봇에게 '닉네임 변경' 권한이 없을 수 있습니다.)",
                    ephemeral=True
                )
            except discord.HTTPException as e:
                await interaction.response.send_message(f"❌ 닉네임 변경 중 오류가 발생했습니다: {e}", ephemeral=True)
        finally:
            if interaction.response.is_done():
                await asyncio.sleep(3)
                try:
                    await interaction.delete_original_response()
                except Exception:
                    pass

    @discord.ui.button(label="대기", style=discord.ButtonStyle.success, custom_id="status_nick_wait")
    async def wait_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        try:
            member = interaction.user
            if not isinstance(member, discord.Member):
                await interaction.response.send_message("❌ 이 버튼은 서버 내에서만 사용할 수 있습니다.", ephemeral=True)
                return

            current_name = member.nick or member.name
            clean_name = current_name
            # 기존 접두사 제거
            prefixes = ["[관전] ", "[대기] ", "[관전]", "[대기]"]
            for prefix in prefixes:
                if clean_name.startswith(prefix):
                    clean_name = clean_name[len(prefix):]
                    break
            clean_name = clean_name.strip()
            new_nick = f"[대기] {clean_name}"[:32]

            try:
                await member.edit(nick=new_nick)
                await interaction.response.send_message(f"✅ 닉네임이 **{new_nick}**(으)로 변경되었습니다.", ephemeral=True)
            except discord.Forbidden:
                await interaction.response.send_message(
                    "❌ 닉네임을 변경할 권한이 없습니다.\n"
                    "(봇의 역할 서열이 본인보다 낮거나, 본인이 서버 소유자이거나, 봇에게 '닉네임 변경' 권한이 없을 수 있습니다.)",
                    ephemeral=True
                )
            except discord.HTTPException as e:
                await interaction.response.send_message(f"❌ 닉네임 변경 중 오류가 발생했습니다: {e}", ephemeral=True)
        finally:
            if interaction.response.is_done():
                await asyncio.sleep(3)
                try:
                    await interaction.delete_original_response()
                except Exception:
                    pass

    @discord.ui.button(label="원래대로", style=discord.ButtonStyle.secondary, custom_id="status_nick_reset")
    async def reset_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        try:
            member = interaction.user
            if not isinstance(member, discord.Member):
                await interaction.response.send_message("❌ 이 버튼은 서버 내에서만 사용할 수 있습니다.", ephemeral=True)
                return

            current_name = member.nick or member.name
            clean_name = current_name
            # 기존 접두사 제거
            prefixes = ["[관전] ", "[대기] ", "[관전]", "[대기]"]
            for prefix in prefixes:
                if clean_name.startswith(prefix):
                    clean_name = clean_name[len(prefix):]
                    break
            clean_name = clean_name.strip()

            # 닉네임을 원래대로 돌리려면, 닉네임이 member.name과 같으면 nick=None을 설정하는 것이 좋음
            new_nick = None if clean_name == member.name else clean_name
            if new_nick is not None:
                new_nick = new_nick[:32]

            try:
                await member.edit(nick=new_nick)
                if new_nick:
                    await interaction.response.send_message(f"✅ 닉네임이 원래대로 (**{new_nick}**) 변경되었습니다.", ephemeral=True)
                else:
                    await interaction.response.send_message("✅ 원래 닉네임(디스코드 기본 이름)으로 복원되었습니다.", ephemeral=True)
            except discord.Forbidden:
                await interaction.response.send_message(
                    "❌ 닉네임을 변경할 권한이 없습니다.\n"
                    "(봇의 역할 서열이 본인보다 낮거나, 본인이 서버 소유자이거나, 봇에게 '닉네임 변경' 권한이 없을 수 있습니다.)",
                    ephemeral=True
                )
            except discord.HTTPException as e:
                await interaction.response.send_message(f"❌ 닉네임 변경 중 오류가 발생했습니다: {e}", ephemeral=True)
        finally:
            if interaction.response.is_done():
                await asyncio.sleep(3)
                try:
                    await interaction.delete_original_response()
                except Exception:
                    pass


def get_seconds_until_next_reset():
    now = datetime.datetime.now()
    # 오늘 오전 6시 설정
    reset_time = now.replace(hour=6, minute=0, second=0, microsecond=0)
    # 현재 시간이 이미 오늘 오전 6시 이후라면, 다음 초기화는 내일 오전 6시임
    if now >= reset_time:
        reset_time += datetime.timedelta(days=1)
    return (reset_time - now).total_seconds()

async def daily_reset_task():
    await bot.wait_until_ready()
    while not bot.is_closed():
        seconds = get_seconds_until_next_reset()
        print(f"⏰ 다음 오전 06시 데이터 분할 및 재시작 대기 시간: {seconds}초")
        await asyncio.sleep(seconds)
        
        try:
            # 현재 음성 채널에 남아 있는 유저들의 누적 시간을 어제 날짜로 기록하고 시작 시점을 오전 6시로 갱신
            now = time.time()
            # 6시 정각 리셋이므로, 이전 날짜(6시간 1분 전)를 기준으로 저장
            target_date = (datetime.datetime.now() - datetime.timedelta(hours=6, minutes=1)).strftime("%Y-%m-%d")
            
            for uid, join_time in list(active_sessions.items()):
                duration = int(now - join_time)
                if duration > 0:
                    add_voice_time(uid, target_date, duration)
                active_sessions[uid] = now
            print(f"📅 오전 06시 정각: {target_date} 기준 음성 채널 이용 시간 데이터 정리가 완료되었습니다.")
            
            # 오전 6시 자동 재시작 진행
            print("🔄 오전 06시 자동 재시작을 진행합니다...")
            await bot.close()
            os.execv(sys.executable, [sys.executable] + sys.argv)
        except Exception as e:
            print(f"❌ 일일 데이터 정리 및 재시작 중 오류 발생: {e}")
            
        await asyncio.sleep(10)


# =========================
# 음성 XP 분당 적립 루프
# =========================
@tasks.loop(minutes=1)
async def voice_xp_loop():
    user_ids = list(active_sessions.keys())
    if not user_ids:
        return
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        p = "%s" if DATABASE_URL else "?"
        
        # 1. 유저 데이터 존재 보장
        for uid in user_ids:
            if DATABASE_URL:
                cursor.execute("INSERT INTO users(user_id) VALUES (%s) ON CONFLICT (user_id) DO NOTHING", (uid,))
            else:
                cursor.execute("INSERT OR IGNORE INTO users(user_id) VALUES (?)", (uid,))
        conn.commit()
        
        # 2. XP, 부스터, 음성 시간 확인 및 일괄 업데이트
        placeholders = ", ".join([p] * len(user_ids))
        cursor.execute(f"SELECT user_id, xp, booster_until, voice_minutes FROM users WHERE user_id IN ({placeholders})", tuple(user_ids))
        rows = cursor.fetchall()
        user_data = {row[0]: {"xp": row[1], "booster_until": row[2], "voice_minutes": row[3]} for row in rows}
        
        now = int(time.time())
        today_str = get_current_date()
        for uid in user_ids:
            data = user_data.get(uid, {"xp": 0, "booster_until": 0, "voice_minutes": 0})
            old_xp = data["xp"]
            booster_until = data["booster_until"]
            old_voice_mins = data["voice_minutes"] if data["voice_minutes"] is not None else 0
            
            is_booster_active = booster_until > now
            # 기본 분당 1에서 5로 변경 (부스터 2배 시 10)
            xp_to_add = 10 if is_booster_active else 5
            coin_to_add = 20  # 분당 20 코인 기본 지급
            
            new_xp = old_xp + xp_to_add
            
            cursor.execute(
                f"UPDATE users SET xp = xp + {p}, coin = coin + {p}, voice_minutes = voice_minutes + 1 WHERE user_id = {p}",
                (xp_to_add, coin_to_add, uid)
            )
            
            # 일일 퀘스트 진행도 적립
            if DATABASE_URL:
                cursor.execute("""
                    INSERT INTO user_quests (user_id, quest_id, progress, claimed, quest_date)
                    VALUES (%s, 'voice_30m', 1, 0, %s)
                    ON CONFLICT (user_id, quest_id, quest_date)
                    DO UPDATE SET progress = user_quests.progress + EXCLUDED.progress
                """, (uid, today_str))
            else:
                cursor.execute("""
                    INSERT INTO user_quests (user_id, quest_id, progress, claimed, quest_date)
                    VALUES (?, 'voice_30m', 1, 0, ?)
                    ON CONFLICT (user_id, quest_id, quest_date)
                    DO UPDATE SET progress = user_quests.progress + EXCLUDED.progress
                """, (uid, today_str))
            
            check_and_grant_level_rewards(cursor, p, uid, old_xp, new_xp)
            
        conn.commit()
        cursor.close()
        conn.close()
        print(f"🎙️ [시즌패스] 음성 활성 유저 {len(user_ids)}명 XP 지급 및 레벨업 체크 완료")
    except Exception as e:
        print(f"❌ voice_xp_loop 오류: {e}")

@bot.event
async def on_ready():
    print(f"✅ 로그인 성공: {bot.user.name} ({bot.user.id})")
    
    # 1. 영구 뷰 등록
    bot.add_view(VoiceUsagePanel())
    bot.add_view(PassPanelView())
    bot.add_view(ShopPanelView())
    bot.add_view(StatusNicknameView())
    
    # 2. 지정된 채널에 패널 메시지가 있는지 확인 및 자동 복구/생성
    target_channel_id = 1513160056214913144
    channel = bot.get_channel(target_channel_id)
    if channel:
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            p = "%s" if DATABASE_URL else "?"
            cursor.execute(f"SELECT message_id FROM voice_panel WHERE channel_id = {p}", (target_channel_id,))
            row = cursor.fetchone()
            panel_msg_id = row[0] if row else None
            
            panel_exists = False
            if panel_msg_id:
                try:
                    await channel.fetch_message(panel_msg_id)
                    panel_exists = True
                except discord.NotFound:
                    # 기존 메시지가 삭제됨
                    cursor.execute(f"DELETE FROM voice_panel WHERE channel_id = {p}", (target_channel_id,))
                    conn.commit()
            
            if not panel_exists:
                embed = discord.Embed(
                    title="📊 음성 채널 이용 시간 조회",
                    description="아래 버튼을 누르면 오늘의 랭킹, 누적 전체 랭킹 또는 특정 날짜의 랭킹을 실시간으로 확인할 수 있습니다.",
                    color=discord.Color.blurple()
                )
                msg = await channel.send(embed=embed, view=VoiceUsagePanel())
                
                # SQLite와 PostgreSQL 공용 ON CONFLICT 문법
                query = f"""
                INSERT INTO voice_panel (channel_id, message_id)
                VALUES ({p}, {p})
                ON CONFLICT (channel_id)
                DO UPDATE SET message_id = EXCLUDED.message_id
                """
                cursor.execute(query, (target_channel_id, msg.id))
                conn.commit()
                print(f"📊 이용 시간 조회 패널 메시지 생성 완료 (ID: {msg.id})")
            else:
                print("📊 이용 시간 조회 패널 메시지가 이미 존재합니다.")
            cursor.close()
            conn.close()
        except Exception as db_err:
            print(f"❌ 패널 메시지 확인/생성 중 DB 오류 발생: {db_err}")
    else:
        print(f"❌ 오류: 패널 채널 ID {target_channel_id}를 찾을 수 없습니다.")

    # 3. 현재 음성 채널에 있는 유저들 세션 초기화
    global active_sessions
    now = time.time()
    active_sessions.clear()
    voice_user_count = 0
    for guild in bot.guilds:
        for voice_channel in guild.voice_channels:
            for member in voice_channel.members:
                if not member.bot:
                    active_sessions[member.id] = now
                    voice_user_count += 1
    print(f"🎙️ 현재 음성 채널에 접속 중인 유저 {voice_user_count}명 세션 등록 완료")
    
    # 4. 일일 데이터 정렬/분할 스케줄러 태스크 시작
    asyncio.create_task(daily_reset_task())
    print("⏰ 일일 음성 채널 데이터 정리 태스크 시작 완료")
    
    # 5. 시즌 패스 XP 적립 루프 시작
    if not voice_xp_loop.is_running():
        voice_xp_loop.start()
        print("⏰ 시즌 패스 음성 XP 적립 루프 시작 완료")

    # 6. 로또 과거 당첨 번호 비동기 캐싱 및 정산 루프 시작
    asyncio.create_task(sync_historical_lotto_data())
    if not lotto_check_loop.is_running():
        lotto_check_loop.start()
        print("⏰ 주간 로또 당첨 결과 자동 정산 루프 시작 완료")

    try:
        # 기존에 복사 등록되어 중복 노출을 유발하던 길드 명령어들을 삭제합니다.
        for guild in bot.guilds:
            bot.tree.clear_commands(guild=guild)
            await bot.tree.sync(guild=guild)
        # 이제 글로벌 명령어로만 1개 노출되도록 동기화합니다.
        await bot.tree.sync()
        print("✅ 글로벌 슬래시 명령어 동기화 및 길드 중복 제거 완료")
    except Exception as e:
        print(f"❌ 슬래시 명령어 동기화 오류: {e}")


# =========================
# 관리자 명령어 (시즌 패스)
# =========================
@bot.tree.command(name="패스패널생성", description="HEAVEN 시즌 패스 버튼 패널을 생성합니다.")
@app_commands.default_permissions(administrator=True)
async def create_pass_panel(interaction: discord.Interaction):
    embed = discord.Embed(
        title="🎫 HEAVEN 시즌 패스",
        description=(
            "음성 채널에 참여하여 패스 레벨을 올리고 풍성한 보상을 획득하세요!\n\n"
            "**💡 획득 방식**\n"
            "🎤 **음성 채널 참여:** 1분당 **5 XP** (부스터 적용 시 **10 XP**) & 💰 **20 코인** 지급\n\n"
            "📦 아래 버튼을 눌러 내 시즌 패스 정보를 확인하거나 상점을 이용하실 수 있습니다."
        ),
        color=0x9b59b6
    )
    await interaction.response.send_message(embed=embed, view=PassPanelView())

@bot.tree.command(name="재화지급", description="관리자용 재화 지급")
@app_commands.default_permissions(administrator=True)
async def give_coin(interaction: discord.Interaction, member: discord.Member, amount: int):
    add_coin(member.id, amount)
    await interaction.response.send_message(f"✅ {member.mention}에게 재화 {amount:,} 지급 완료.", ephemeral=True)

@bot.tree.command(name="상자지급", description="관리자용 상자 지급")
@app_commands.default_permissions(administrator=True)
@app_commands.choices(box_type=[
    app_commands.Choice(name="📦 랜덤 상자", value="random_box"),
    app_commands.Choice(name="🎁 프리미엄 상자", value="premium_box"),
    app_commands.Choice(name="👑 잭팟 상자", value="jackpot_box")
])
async def give_box(interaction: discord.Interaction, member: discord.Member, box_type: str, amount: int):
    add_item(member.id, box_type, amount)
    box_names = {
        "random_box": "랜덤 상자",
        "premium_box": "프리미엄 상자",
        "jackpot_box": "잭팟 상자"
    }
    await interaction.response.send_message(f"✅ {member.mention}에게 {box_names[box_type]} {amount}개 지급 완료.", ephemeral=True)

@bot.tree.command(name="xp지급", description="관리자용 XP 지급")
@app_commands.default_permissions(administrator=True)
async def give_xp(interaction: discord.Interaction, member: discord.Member, amount: int):
    rewards = add_xp(member.id, amount)
    msg = f"✅ {member.mention}에게 XP {amount:,} 지급 완료."
    if rewards:
        msg += f"\n🎁 지급 과정에서 레벨업 보상 획득: {', '.join(rewards)}"
    await interaction.response.send_message(msg, ephemeral=True)


@bot.tree.command(name="로또", description="로또 6/45 추천 번호를 생성합니다.")
@app_commands.describe(
    수량="생성할 로또 게임 수량 (1~20개, 기본값: 1)",
    고정수="반드시 포함할 번호 (쉼표 또는 공백 구분, 예: 3,14)",
    제외수="생성에서 제외할 번호 (쉼표 또는 공백 구분, 예: 4 45)",
    필터="적용할 AI 패턴 필터 (기본값: 기본 균형형)"
)
@app_commands.choices(필터=[
    app_commands.Choice(name="⚖️ 기본 균형형", value="balanced"),
    app_commands.Choice(name="🚫 연속수 배제형", value="no_consecutive"),
    app_commands.Choice(name="🔴 홀수 강조형", value="odd_heavy"),
    app_commands.Choice(name="🔵 짝수 강조형", value="even_heavy"),
    app_commands.Choice(name="🟢 고수 강조형", value="high_heavy"),
    app_commands.Choice(name="🟡 소수 강조형", value="low_heavy"),
    app_commands.Choice(name="🎰 순수 무작위형", value="random")
])
async def lotto_recommend(
    interaction: discord.Interaction,
    수량: int = 1,
    고정수: str = None,
    제외수: str = None,
    필터: str = "balanced"
):
    # 1. 수량 유효성 검사 (1~20)
    if 수량 < 1 or 수량 > 20:
        await interaction.response.send_message("⚠️ 생성할 수량은 1개에서 20개까지만 선택 가능합니다.", ephemeral=True)
        return

    # 2. 고정수/제외수 파싱 및 검사
    fixed_nums, fixed_err = parse_number_list(고정수)
    if fixed_err:
        await interaction.response.send_message(f"❌ 고정수 입력 오류: {fixed_err}", ephemeral=True)
        return
        
    excluded_nums, excluded_err = parse_number_list(제외수)
    if excluded_err:
        await interaction.response.send_message(f"❌ 제외수 입력 오류: {excluded_err}", ephemeral=True)
        return

    # 3. 고정수와 제외수 중복 및 개수 검사
    overlap = fixed_nums.intersection(excluded_nums)
    if overlap:
        await interaction.response.send_message(
            f"❌ 오류: 고정수와 제외수에 동시에 지정된 번호가 있습니다: `{' '.join(str(n) for n in overlap)}`",
            ephemeral=True
        )
        return
        
    if len(fixed_nums) > 5:
        await interaction.response.send_message("❌ 고정수는 최대 5개까지만 설정할 수 있습니다.", ephemeral=True)
        return
        
    if len(fixed_nums) + len(excluded_nums) > 40:
        await interaction.response.send_message("❌ 고정수와 제외수의 총합이 너무 많아 로또 생성 조건을 충족할 수 없습니다.", ephemeral=True)
        return

    # 4. View 생성 및 Embed 출력
    view = LottoRecommendView(
        user_id=interaction.user.id,
        count=수량,
        fixed_nums=fixed_nums,
        excluded_nums=excluded_nums,
        pattern=필터
    )
    
    embed = view.create_embed()
    await interaction.response.send_message(embed=embed, view=view)


@bot.tree.command(name="로또분석", description="입력한 번호 조합의 최근 100회차 통계 및 분석을 제공합니다.")
@app_commands.describe(번호="분석할 로또 번호 6개 (쉼표 또는 공백 구분, 예: 3, 14, 25, 36, 42, 45)")
async def lotto_analyze(interaction: discord.Interaction, 번호: str):
    nums, err = parse_number_list(번호)
    if err:
        await interaction.response.send_message(f"❌ 입력 오류: {err}", ephemeral=True)
        return
        
    if len(nums) != 6:
        await interaction.response.send_message("❌ 로또 번호는 정확히 6개 입력하셔야 합니다.", ephemeral=True)
        return

    sorted_nums = sorted(list(nums))
    frequencies, last_seen, actual_count = get_lotto_stats_from_db()
    current_round = get_current_lotto_round()
    
    def get_emoji(num):
        if 1 <= num <= 10: return "🟡"
        elif 11 <= num <= 20: return "🔵"
        elif 21 <= num <= 30: return "🔴"
        elif 31 <= num <= 40: return "⚫"
        else: return "🟢"

    formatted_numbers = "  ".join([f"{get_emoji(num)} `{num:02d}`" for num in sorted_nums])
    
    total_sum = sum(sorted_nums)
    odds = sum(1 for n in sorted_nums if n % 2 != 0)
    evens = 6 - odds
    lows = sum(1 for n in sorted_nums if n <= 22)
    highs = 6 - lows
    
    diffs = set()
    for i in range(len(sorted_nums)):
        for j in range(i + 1, len(sorted_nums)):
            diffs.add(sorted_nums[j] - sorted_nums[i])
    ac_val = len(diffs) - (6 - 1)
    
    consec_pairs = []
    for i in range(len(sorted_nums) - 1):
        if sorted_nums[i+1] - sorted_nums[i] == 1:
            consec_pairs.append(f"({sorted_nums[i]}, {sorted_nums[i+1]})")
    consec_text = ", ".join(consec_pairs) if consec_pairs else "없음"
    
    freq_details = []
    cold_details = []
    for num in sorted_nums:
        freq = frequencies.get(num, 0)
        last_rd = last_seen.get(num, 0)
        cold_weeks = current_round - last_rd if last_rd > 0 else 100
        cold_text = f"{cold_weeks}주" if cold_weeks < 100 else "100주+"
        
        freq_details.append(f"`{num:02d}`({freq}회)")
        cold_details.append(f"`{num:02d}`({cold_text})")
    
    avg_freq = sum(frequencies.get(num, 0) for num in sorted_nums) / 6
    avg_cold = sum((current_round - last_seen.get(num, 0)) if last_seen.get(num, 0) > 0 else 100 for num in sorted_nums) / 6
    
    embed = discord.Embed(
        title="🔍 로또 번호 심층 분석 결과",
        description=f"입력하신 조합의 과거 **최근 100회차** 데이터를 기반으로 분석한 결과입니다.",
        color=0x00F0FF  # 네온 블루
    )
    
    embed.add_field(name="🎫 분석 대상 번호", value=formatted_numbers, inline=False)
    
    analysis_text = (
        f"▪️ **총합:** `{total_sum}` {'(균형: 100~180)' if 100 <= total_sum <= 180 else '(비균형)'}\n"
        f"▪️ **홀짝 비율:** `{odds}:{evens}`\n"
        f"▪️ **고저 비율:** `{lows}:{highs}` (Low: 1~22, High: 23~45)\n"
        f"▪️ **산술 복잡도 (AC값):** `{ac_val}` (5 이상 권장)\n"
        f"▪️ **연속 번호 쌍:** `{consec_text}`\n"
        f"▪️ **평균 출현 빈도:** `{avg_freq:.1f}회` (11~15회 권장, 최근 100회 기준)\n"
        f"▪️ **평균 미출현 기간:** `{avg_cold:.1f}주` (6~10주 권장)"
    )
    embed.add_field(name="📊 패턴 및 종합 통계 분석", value=analysis_text, inline=False)
    
    detail_stat_text = (
        f"▪️ **번호별 출현 빈도:** {', '.join(freq_details)}\n"
        f"▪️ **번호별 미출현 기간:** {', '.join(cold_details)}"
    )
    embed.add_field(name="🔍 번호별 심층 통계 (최근 100회차)", value=detail_stat_text, inline=False)
    
    embed.set_footer(text="HEAVEN AI 로또 연구소")
    
    await interaction.response.send_message(embed=embed)


def get_saved_lotto_tickets_embed(user_id):
    current_round = get_current_lotto_round()
    
    conn = get_db_connection()
    cursor = conn.cursor()
    p = "%s" if DATABASE_URL else "?"
    
    # 1. Fetch un-drawn tickets (is_checked = 0)
    cursor.execute(
        f"SELECT round_no, numbers, created_at FROM lotto_tickets WHERE user_id = {p} AND is_checked = 0 ORDER BY round_no ASC, id ASC",
        (user_id,)
    )
    undrawn_rows = cursor.fetchall()
    
    # 2. Fetch drawn tickets (is_checked = 1, limit to 5)
    cursor.execute(
        f"SELECT round_no, numbers, match_count, prize_rank, created_at FROM lotto_tickets WHERE user_id = {p} AND is_checked = 1 ORDER BY round_no DESC, id DESC LIMIT 5",
        (user_id,)
    )
    drawn_rows = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    embed = discord.Embed(
        title="💾 나의 저장된 로또 번호 내역",
        description="HEAVEN AI 로또 연구소에서 저장한 로또 번호 목록입니다.",
        color=0xFF007F
    )
    
    def get_emoji(num):
        if 1 <= num <= 10: return "🟡"
        elif 11 <= num <= 20: return "🔵"
        elif 21 <= num <= 30: return "🔴"
        elif 31 <= num <= 40: return "⚫"
        else: return "🟢"
        
    def format_numbers(nums_str):
        nums = [int(n) for n in nums_str.split(",")]
        return " ".join(f"{get_emoji(n)}`{n:02d}`" for n in sorted(nums))

    # Active/Un-drawn tickets field
    if undrawn_rows:
        active_by_round = {}
        for round_no, numbers, created_at in undrawn_rows:
            active_by_round.setdefault(round_no, []).append((numbers, created_at))
            
        for round_no, tickets in active_by_round.items():
            ticket_texts = []
            for i, (numbers, created_at) in enumerate(tickets, 1):
                date_part = created_at.split()[0] if created_at else ""
                ticket_texts.append(f"`{i:02d}`. {format_numbers(numbers)} `({date_part})`")
            embed.add_field(
                name=f"🎟️ 제 {round_no}회차 대기 중 조합 ({len(tickets)}개)",
                value="\n".join(ticket_texts),
                inline=False
            )
    else:
        embed.add_field(
            name="🎟️ 미추첨 대기 중 조합",
            value="현재 저장된 대기 중인 번호 조합이 없습니다.\n`/로또` 또는 `!로또` 명령어로 번호를 생성하고 저장해보세요!",
            inline=False
        )
        
    # Drawn tickets field
    if drawn_rows:
        drawn_texts = []
        rank_names = {
            1: "🥇 1등 당첨",
            2: "🥈 2등 당첨",
            3: "🥉 3등 당첨",
            4: "💎 4등 당첨",
            5: "🍀 5등 당첨",
            0: "❌ 낙첨"
        }
        for round_no, numbers, match_count, prize_rank, created_at in drawn_rows:
            rank_text = rank_names.get(prize_rank, "❌ 낙첨")
            drawn_texts.append(
                f"**제 {round_no}회차** | {format_numbers(numbers)}\n"
                f"└ 정산: `{rank_text}` (일치 개수: {match_count}개)"
            )
        embed.add_field(
            name="📊 최근 당첨 정산 내역 (최근 최대 5개)",
            value="\n\n".join(drawn_texts),
            inline=False
        )
    else:
        embed.add_field(
            name="📊 최근 당첨 정산 내역",
            value="최근 정산된 로또 내역이 없습니다.",
            inline=False
        )
        
    embed.set_footer(text="HEAVEN AI 로또 연구소")
    return embed


@bot.tree.command(name="상태패널생성", description="관전/대기 닉네임 상태설정 패널을 생성합니다.")
@app_commands.default_permissions(administrator=True)
async def create_status_panel(interaction: discord.Interaction):
    embed = discord.Embed(
        title="🏷️ 관전대기변경",
        description="아래 버튼을 클릭하여 닉네임 접두사를 변경할 수 있습니다.\n\n"
                    "📋 **사용 가능한 접두사**\n"
                    "• **관전** - `[관전] [닉네임]` 형태로 변경\n"
                    "• **대기** - `[대기] [닉네임]` 형태로 변경\n"
                    "• **원래대로** - 원래 닉네임으로 복원",
        color=discord.Color.blue()
    )
    await interaction.response.send_message("✅ 상태 변경 패널을 생성했습니다.", ephemeral=True)
    await interaction.channel.send(embed=embed, view=StatusNicknameView())


@bot.tree.command(name="팀나누기", description="대기방 채널의 유저들을 팀으로 나누어 이동시킵니다.")
@app_commands.describe(team_size="한 팀당 인원수 (기본값: 5)")
async def split_teams(interaction: discord.Interaction, team_size: int = 5):
    if team_size <= 0:
        await interaction.response.send_message("❌ 팀 인원수는 1명 이상이어야 합니다.", ephemeral=True)
        return
        
    hub_channel_id = 1532691400230047805
    category_id = 1532692129569046559
    
    guild = interaction.guild
    hub_channel = guild.get_channel(hub_channel_id)
    category = guild.get_channel(category_id)
    
    if not hub_channel or not isinstance(hub_channel, discord.VoiceChannel):
        await interaction.response.send_message("❌ 대상 대기방 음성 채널을 찾을 수 없습니다.", ephemeral=True)
        return
        
    if not category or not isinstance(category, discord.CategoryChannel):
        await interaction.response.send_message("❌ 대상 카테고리를 찾을 수 없습니다.", ephemeral=True)
        return
        
    all_members = [m for m in hub_channel.members if not m.bot]
    active_members = [m for m in all_members if not m.display_name.startswith("[관전]")]
    spectator_members = [m for m in all_members if m.display_name.startswith("[관전]")]
    
    if not active_members:
        await interaction.response.send_message("❌ 대상 대기방 채널에 플레이어가 없습니다. (관전자만 있는 경우 팀을 나눌 수 없습니다.)", ephemeral=True)
        return
        
    await interaction.response.defer(ephemeral=True)
    
    # 섞기 및 플레이어 분배
    random.shuffle(active_members)
    chunks = [active_members[i:i + team_size] for i in range(0, len(active_members), team_size)]
    
    import string
    created_channels = []
    failed_moves = 0
    success_moves = 0
    
    for chunk in chunks:
        # 비어 있는 알파벳 팀 채널명 찾기
        existing_names = [c.name for c in category.voice_channels]
        new_name = None
        for char in string.ascii_uppercase:
            candidate = f"{char}팀"
            if candidate not in existing_names:
                new_name = candidate
                break
        if not new_name:
            i = 2
            while True:
                candidate = f"A팀 {i}"
                if candidate not in existing_names:
                    new_name = candidate
                    break
                i += 1
                
        try:
            # 채널 생성
            new_channel = await guild.create_voice_channel(name=new_name, category=category)
            created_channels.append(new_channel)
            
            # 플레이어 이동
            for member in chunk:
                try:
                    await member.move_to(new_channel)
                    success_moves += 1
                except Exception as move_err:
                    print(f"❌ 멤버 {member.name} 이동 실패: {move_err}")
                    failed_moves += 1
        except discord.Forbidden:
            await interaction.followup.send("❌ 채널 생성 권한이 부족합니다.", ephemeral=True)
            return
        except Exception as create_err:
            await interaction.followup.send(f"❌ 채널 생성 중 오류 발생: {create_err}", ephemeral=True)
            return

    # 관전자 이동 처리 (생성된 채널 중 랜덤 배정)
    if spectator_members and created_channels:
        for spectator in spectator_members:
            target_channel = random.choice(created_channels)
            try:
                await spectator.move_to(target_channel)
                success_moves += 1
            except Exception as move_err:
                print(f"❌ 관전자 {spectator.name} 이동 실패: {move_err}")
                failed_moves += 1

    result_msg = f"✅ 플레이어 {len(active_members)}명을 {team_size}명씩 나누어 {len(chunks)}개의 팀 채널을 생성했습니다.\n"
    if spectator_members:
        result_msg += f"👁️ 관전자 {len(spectator_members)}명도 각 팀 채널에 랜덤하게 배정되었습니다.\n"
    result_msg += f"🔊 생성된 채널: {', '.join([c.name for c in created_channels])}\n"
    result_msg += f"👥 이동 완료: {success_moves}명"
    if failed_moves > 0:
        result_msg += f" (실패: {failed_moves}명)"
        
    await interaction.followup.send(result_msg, ephemeral=True)




    


@bot.tree.command(name="로또조회", description="저장된 나의 로또 번호 내역을 조회합니다.")
async def lotto_lookup(interaction: discord.Interaction):
    embed = get_saved_lotto_tickets_embed(interaction.user.id)
    await interaction.response.send_message(embed=embed, ephemeral=True)

# ------------------ 관리자용 신규 명령어 및 UI (안전장치 추가됨) ------------------

class MultiUserKickConfirmView(discord.ui.View):
    def __init__(self, targets, admin_user):
        super().__init__(timeout=60)
        self.targets = targets
        self.admin_user = admin_user

    @discord.ui.button(label="추방하기", style=discord.ButtonStyle.danger)
    async def confirm_kick(self, interaction: discord.Interaction, button: discord.ui.Button):
        if interaction.user.id != self.admin_user.id:
            await interaction.response.send_message("❌ 이 버튼은 명령어를 실행한 관리자만 사용할 수 있습니다.", ephemeral=True)
            return

        import asyncio
        await interaction.response.defer(ephemeral=True)
        
        # 버튼을 비활성화하고 작업 시작을 알린다.
        for child in self.children:
            child.disabled = True
        await interaction.edit_original_response(
            content=f"⏳ 총 {len(self.targets)}명의 유저를 추방하는 중입니다. 잠시만 기다려주세요...\n(디스코드 제한 방지를 위해 1초 간격으로 처리됩니다.)",
            embed=None,
            view=self
        )

        success_list = []
        fail_list = []

        for member in self.targets:
            try:
                # 실제로 추방 실행
                await member.kick(reason=f"관리자 {self.admin_user.name}에 의한 일괄 추방")
                success_list.append(f"{member.name} ({member.id})")
                await asyncio.sleep(1.0) # 안전장치: Rate Limit 방어
            except discord.Forbidden:
                fail_list.append(f"{member.name} ({member.id}) - 권한 부족")
            except Exception as e:
                fail_list.append(f"{member.name} ({member.id}) - {str(e)}")

        result_embed = discord.Embed(
            title="✅ 일괄 추방 완료",
            description=f"추방 요청 {len(self.targets)}건 중 성공 {len(success_list)}건, 실패 {len(fail_list)}건",
            color=discord.Color.orange(),
            timestamp=datetime.datetime.now()
        )

        if success_list:
            success_str = "\n".join(success_list[:15])
            if len(success_list) > 15:
                success_str += f"\n... 외 {len(success_list) - 15}명"
            result_embed.add_field(name="🟢 성공 유저 목록", value=f"```\n{success_str}\n```", inline=False)

        if fail_list:
            fail_str = "\n".join(fail_list[:15])
            if len(fail_list) > 15:
                fail_str += f"\n... 외 {len(fail_list) - 15}명"
            result_embed.add_field(name="🔴 실패 유저 목록 (권한 부족 등)", value=f"```\n{fail_str}\n```", inline=False)

        # 모든 버튼 비활성화
        for child in self.children:
            child.disabled = True

        await interaction.followup.edit_message(message_id=interaction.message.id, embed=result_embed, view=None)

    @discord.ui.button(label="취소", style=discord.ButtonStyle.secondary)
    async def cancel_kick(self, interaction: discord.Interaction, button: discord.ui.Button):
        if interaction.user.id != self.admin_user.id:
            await interaction.response.send_message("❌ 이 버튼은 명령어를 실행한 관리자만 사용할 수 있습니다.", ephemeral=True)
            return

        for child in self.children:
            child.disabled = True

        await interaction.response.edit_message(content="🛑 추방 요청이 취소되었습니다.", embed=None, view=None)
class PaginationUserSelect(discord.ui.Select):
    def __init__(self, members, page, parent_view):
        self.parent_view = parent_view
        self.page_members = members[page*25 : (page+1)*25]
        
        options = []
        for m in self.page_members:
            label = f"{m.name}"
            if m.nick and m.nick != m.name:
                label += f" ({m.nick})"
            label = label[:100]
            
            is_selected = m.id in self.parent_view.selected_ids
            prefix = "☑️ " if is_selected else "⬜ "
            
            options.append(discord.SelectOption(
                label=f"{prefix}{label}",
                value=str(m.id),
                description=f"ID: {m.id}"
            ))
            
        if not options:
            options = [discord.SelectOption(label="멤버 없음", value="none")]
            
        super().__init__(
            placeholder=f"멤버를 선택해 주세요 (페이지 {page+1}/{(len(members)-1)//25 + 1})",
            min_values=1 if options[0].value != "none" else 0,
            max_values=len(options) if options[0].value != "none" else 0,
            options=options
        )
        
    async def callback(self, interaction: discord.Interaction):
        if interaction.user.id != self.parent_view.admin_user.id:
            await interaction.response.send_message("❌ 이 메뉴는 명령어를 실행한 관리자만 사용할 수 있습니다.", ephemeral=True)
            return
            
        if self.values and self.values[0] != "none":
            for val in self.values:
                uid = int(val)
                if uid in self.parent_view.selected_ids:
                    self.parent_view.selected_ids.remove(uid)
                else:
                    self.parent_view.selected_ids.add(uid)
            
            await self.parent_view.update_view(interaction)


class MultiUserKickView(discord.ui.View):
    def __init__(self, admin_user, guild):
        super().__init__(timeout=300)
        self.admin_user = admin_user
        self.guild = guild
        self.all_members = sorted([m for m in guild.members if not m.bot], key=lambda x: x.display_name.lower())
        self.selected_ids = set()
        self.current_page = 0
        self.total_pages = (len(self.all_members) - 1) // 25 + 1 if self.all_members else 1
        
        self.setup_components()
        
    def setup_components(self):
        self.clear_items()
        self.add_item(PaginationUserSelect(self.all_members, self.current_page, self))
        
        prev_btn = discord.ui.Button(label="◀ 이전", style=discord.ButtonStyle.secondary, disabled=(self.current_page == 0))
        prev_btn.callback = self.prev_page
        self.add_item(prev_btn)
        
        page_btn = discord.ui.Button(label=f"{self.current_page + 1} / {self.total_pages}", style=discord.ButtonStyle.secondary, disabled=True)
        self.add_item(page_btn)
        
        next_btn = discord.ui.Button(label="다음 ▶", style=discord.ButtonStyle.secondary, disabled=(self.current_page >= self.total_pages - 1))
        next_btn.callback = self.next_page
        self.add_item(next_btn)
        
        confirm_btn = discord.ui.Button(
            label=f"⚠️ {len(self.selected_ids)}명 추방하기", 
            style=discord.ButtonStyle.danger, 
            disabled=(len(self.selected_ids) == 0)
        )
        confirm_btn.callback = self.go_to_confirm
        self.add_item(confirm_btn)
        
        cancel_btn = discord.ui.Button(label="🛑 취소", style=discord.ButtonStyle.secondary)
        cancel_btn.callback = self.cancel_action
        self.add_item(cancel_btn)
        
    async def update_view(self, interaction: discord.Interaction):
        self.setup_components()
        
        embed = discord.Embed(
            title="👤 유저 일괄 관리 패널 (페이지네이션)",
            description="목록에서 추방할 멤버들을 선택해 주세요. 여러 페이지를 넘나들며 선택할 수 있습니다.\n"
                        "선택된 멤버 앞에는 ☑️ 표시가 붙으며, 아래 [추방하기] 버튼에 실시간 반영됩니다.",
            color=discord.Color.blue()
        )
        
        if self.selected_ids:
            mentions = []
            for uid in self.selected_ids:
                m = self.guild.get_member(uid)
                if m:
                    mentions.append(f"{m.mention} ({m.name})")
                else:
                    mentions.append(f"알 수 없는 사용자 (ID: {uid})")
            
            selected_str = ", ".join(mentions)
            if len(selected_str) > 1024:
                selected_str = selected_str[:1000] + " ... 외 다수"
            embed.add_field(name="📍 현재 선택된 멤버 목록", value=selected_str, inline=False)
        else:
            embed.add_field(name="📍 현재 선택된 멤버 목록", value="선택된 유저가 없습니다.", inline=False)
            
        await interaction.response.edit_message(embed=embed, view=self)
        
    async def prev_page(self, interaction: discord.Interaction):
        if interaction.user.id != self.admin_user.id:
            await interaction.response.send_message("❌ 이 버튼은 명령어를 실행한 관리자만 사용할 수 있습니다.", ephemeral=True)
            return
        if self.current_page > 0:
            self.current_page -= 1
            await self.update_view(interaction)
            
    async def next_page(self, interaction: discord.Interaction):
        if interaction.user.id != self.admin_user.id:
            await interaction.response.send_message("❌ 이 버튼은 명령어를 실행한 관리자만 사용할 수 있습니다.", ephemeral=True)
            return
        if self.current_page < self.total_pages - 1:
            self.current_page += 1
            await self.update_view(interaction)
            
    async def go_to_confirm(self, interaction: discord.Interaction):
        if interaction.user.id != self.admin_user.id:
            await interaction.response.send_message("❌ 이 버튼은 명령어를 실행한 관리자만 사용할 수 있습니다.", ephemeral=True)
            return
            
        selected_members = []
        for uid in self.selected_ids:
            m = self.guild.get_member(uid)
            if m:
                selected_members.append(m)
                
        if not selected_members:
            await interaction.response.send_message("❌ 선택된 멤버를 서버에서 찾을 수 없습니다.", ephemeral=True)
            return
            
        embed = discord.Embed(
            title="⚠️ 유저 일괄 추방 확인",
            description=f"정말로 선택한 **{len(selected_members)}명**의 유저를 서버에서 추방하시겠습니까?",
            color=discord.Color.red()
        )
        
        member_list_str = "\n".join([f"• {m.mention} ({m.name} / ID: {m.id})" for m in selected_members])
        if len(member_list_str) > 1024:
            member_list_str = member_list_str[:1000] + "\n... 외 다수"
            
        embed.add_field(name="추방 대상자 목록", value=member_list_str, inline=False)
        embed.set_footer(text="추방된 유저는 서버에서 즉시 내보내집니다.")
        
        confirm_view = MultiUserKickConfirmView(selected_members, self.admin_user)
        await interaction.response.edit_message(content=None, embed=embed, view=confirm_view)
        
    async def cancel_action(self, interaction: discord.Interaction):
        if interaction.user.id != self.admin_user.id:
            await interaction.response.send_message("❌ 이 버튼은 명령어를 실행한 관리자만 사용할 수 있습니다.", ephemeral=True)
            return
        self.clear_items()
        await interaction.response.edit_message(content="🛑 추방 작업이 취소되었습니다.", embed=None, view=None)


@bot.tree.command(name="유저관리", description="[관리자 전용] 유저를 다중 선택하여 일괄 추방(Kick)할 수 있는 UI를 호출합니다.")
@app_commands.checks.has_permissions(kick_members=True)
async def user_management(interaction: discord.Interaction):
    # 본인만 볼 수 있는 비공개 메시지로 유저 관리 패널 전송
    view = MultiUserKickView(admin_user=interaction.user, guild=interaction.guild)
    embed = discord.Embed(
        title="👤 유저 일괄 관리 패널 (페이지네이션)",
        description="목록에서 추방할 멤버들을 선택해 주세요. 여러 페이지를 넘나들며 선택할 수 있습니다.\n"
                    "선택된 멤버 앞에는 ☑️ 표시가 붙으며, 아래 [추방하기] 버튼에 실시간 반영됩니다.",
        color=discord.Color.blue()
    )
    embed.add_field(name="📍 현재 선택된 멤버 목록", value="선택된 유저가 없습니다.", inline=False)
    await interaction.response.send_message(
        embed=embed,
        view=view,
        ephemeral=True
    )



@bot.tree.command(name="유저목록엑셀", description="[관리자 전용] 서버의 모든 멤버 프로필과 DB 데이터를 결합한 엑셀 파일을 추출합니다.")
@app_commands.checks.has_permissions(administrator=True)
async def user_list_excel(interaction: discord.Interaction):
    await interaction.response.defer(ephemeral=True)

    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    try:
        # 1. 엑셀 워크북 초기화
        wb = Workbook()
        ws = wb.active
        ws.title = "서버 유저 목록"

        # 2. 헤더 스타일 정의
        font_bold = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
        font_normal = Font(name="맑은 고딕", size=10)
        fill_header = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")

        border_side = Side(style="thin", color="D9D9D9")
        border_all = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

        headers = [
            "유저 ID", "디스코드 닉네임", "서버 닉네임", "계정 생성일", "서버 가입일", 
            "보유 역할", "XP", "코인", "음성 누적(분)"
        ]

        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = font_bold
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_all

        # 3. DB 데이터 로드
        conn = get_db_connection()
        cursor = conn.cursor()

        db_users = {}
        try:
            cursor.execute("SELECT user_id, xp, coin, voice_minutes FROM users")
            rows = cursor.fetchall()
            for r in rows:
                db_users[r[0]] = {
                    "xp": r[1],
                    "coin": r[2],
                    "voice_minutes": r[3]
                }
        except Exception as db_err:
            print(f"❌ 엑셀용 DB 조회 중 오류: {db_err}")
        finally:
            cursor.close()
            conn.close()

        # 4. 멤버 목록 조회 및 행 작성
        row_count = 2
        
        # 봇 제외한 멤버 데이터 구성
        for member in interaction.guild.members:
            if member.bot:
                continue

            uid = member.id
            name = member.name
            display_name = member.display_name
            created_at = member.created_at.strftime("%Y-%m-%d %H:%M:%S")
            joined_at = member.joined_at.strftime("%Y-%m-%d %H:%M:%S") if member.joined_at else "알 수 없음"

            # 역할 목록 (@everyone 제외)
            roles = [r.name for r in member.roles if r.name != "@everyone"]
            roles_str = ", ".join(roles) if roles else "역할 없음"

            # DB 연동 데이터
            xp = db_users.get(uid, {}).get("xp", 0)
            coin = db_users.get(uid, {}).get("coin", 0)
            voice_min = db_users.get(uid, {}).get("voice_minutes", 0)

            row_data = [
                str(uid), name, display_name, created_at, joined_at,
                roles_str, xp, coin, voice_min
            ]

            ws.append(row_data)

            # 데이터 행 스타일 적용
            for col_num in range(1, len(row_data) + 1):
                cell = ws.cell(row=row_count, column=col_num)
                cell.font = font_normal
                cell.border = border_all

                # 정렬 및 서식
                if col_num in [1, 4, 5]: # ID, 날짜들
                    cell.alignment = align_center
                elif col_num in [7, 8, 9]: # 숫자 데이터
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = "#,##0"
                else:
                    cell.alignment = align_left

            row_count += 1

        # 5. 열 너비 자동 조정
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or "")
                lines = val_str.split('\n')
                for line in lines:
                    kor_count = len(re.findall(r'[ㄱ-ㅎ|ㅏ-ㅣ|가-힣]', line))
                    line_len = len(line) + kor_count
                    if line_len > max_len:
                        max_len = line_len
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

        # 6. 바이트 버퍼에 저장하여 파일 전송
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        file = discord.File(file_stream, filename=f"서버유저목록_{interaction.guild.name}_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx")
        await interaction.followup.send(file=file, content="📊 서버 유저 목록 엑셀 파일 추출이 완료되었습니다.", ephemeral=True)

    except Exception as e:
        print(f"❌ 유저목록엑셀 추출 중 오류: {e}")
        await interaction.followup.send(content=f"❌ 엑셀 추출 중 오류가 발생했습니다: {e}", ephemeral=True)


# 권한 예외 처리 핸들러
@user_management.error
async def user_management_error(interaction: discord.Interaction, error: app_commands.AppCommandError):
    if isinstance(error, app_commands.errors.MissingPermissions):
        await interaction.response.send_message("❌ 이 명령어는 추방(Kick) 권한이 있는 관리자만 사용할 수 있습니다.", ephemeral=True)
    else:
        await interaction.response.send_message("❌ 명령어를 처리하는 중에 오류가 발생했습니다.", ephemeral=True)

@user_list_excel.error
async def user_list_excel_error(interaction: discord.Interaction, error: app_commands.AppCommandError):
    if isinstance(error, app_commands.errors.MissingPermissions):
        await interaction.response.send_message("❌ 이 명령어는 관리자(Administrator) 권한이 있는 유저만 사용할 수 있습니다.", ephemeral=True)
    else:
        await interaction.response.send_message("❌ 명령어를 처리하는 중에 오류가 발생했습니다.", ephemeral=True)


# =========================
# 월간 활동 기록 & 미접속자 명령어
# =========================
@bot.tree.command(name="월간기록", description="매월 1일~말일 음성 활동 기록과 2주 이상 미접속자 현황을 조회합니다.")
@app_commands.describe(조회월="조회할 연월 (예: 2026-09, 미입력 시 이번 달)")
async def monthly_record(interaction: discord.Interaction, 조회월: str = None):
    now = get_kst_now()
    target_year = now.year
    target_month = now.month
    
    if 조회월:
        try:
            cleaned = 조회월.strip().replace(".", "-").replace("/", "-")
            parts = cleaned.split("-")
            if len(parts) == 2:
                target_year = int(parts[0])
                target_month = int(parts[1])
                if not (1 <= target_month <= 12):
                    raise ValueError
            else:
                raise ValueError
        except Exception:
            await interaction.response.send_message("❌ 올바른 연월 형식이 아닙니다. (예: `2026-09` 또는 `2026.09`)", ephemeral=True)
            return

    embed = build_monthly_dashboard_embed(interaction.guild, target_year, target_month)
    view = MonthlyDashboardView(interaction.guild, target_year, target_month)
    await interaction.response.send_message(embed=embed, view=view)


@bot.tree.command(name="미접속자", description="2주(14일) 이상 음성 채널에 접속하지 않은 멤버 목록과 미접속 일수를 조회합니다.")
async def inactive_members_cmd(interaction: discord.Interaction):
    now = get_kst_now()
    inactive_list = get_inactive_members(interaction.guild, days_threshold=14) if interaction.guild else []
    if not inactive_list:
        await interaction.response.send_message("✨ 2주(14일) 이상 미접속 멤버가 없습니다! 모든 멤버가 활동 중입니다.", ephemeral=True)
        return
    view = InactiveMembersView(interaction.guild, now.year, now.month, inactive_list)
    embed = view.create_embed()
    await interaction.response.send_message(embed=embed, view=view)


# 양식 입력 확인 및 역할 제거 이벤트
@bot.event
async def on_message(message):
    if message.author.bot:
        return

    # 특정 채널(ID: 1516049566535909439)의 메시지는 봇이 무시하도록 설정
    if message.channel.id == 1516049566535909439:
        return

    # "로또" 또는 "!로또" 텍스트 명령어 대응
    if message.content.strip() == "로또" or message.content.strip() == "!로또":
        view = LottoRecommendView(
            user_id=message.author.id,
            count=1,
            fixed_nums=set(),
            excluded_nums=set(),
            pattern="balanced"
        )
        embed = view.create_embed()
        await message.channel.send(embed=embed, view=view)
        return

    # "로또조회" 또는 "!로또조회" 또는 "!내로또" 텍스트 명령어 대응
    if message.content.strip() in ["로또조회", "!로또조회", "내로또", "!내로또"]:
        embed = get_saved_lotto_tickets_embed(message.author.id)
        await message.channel.send(embed=embed)
        return

    # "!월간기록" 또는 "!활동기록" 텍스트 명령어 대응
    if message.content.strip() in ["!월간기록", "월간기록", "!활동기록", "활동기록"]:
        now = get_kst_now()
        embed = build_monthly_dashboard_embed(message.guild, now.year, now.month)
        view = MonthlyDashboardView(message.guild, now.year, now.month)
        await message.channel.send(embed=embed, view=view)
        return

    # "!미접속자" 텍스트 명령어 대응
    if message.content.strip() in ["!미접속자", "미접속자"]:
        now = get_kst_now()
        inactive_list = get_inactive_members(message.guild, days_threshold=14) if message.guild else []
        if not inactive_list:
            await message.channel.send("✨ 2주(14일) 이상 미접속 멤버가 없습니다! 모든 멤버가 활동 중입니다.")
            return
        view = InactiveMembersView(message.guild, now.year, now.month, inactive_list)
        embed = view.create_embed()
        await message.channel.send(embed=embed, view=view)
        return

    # "!상태패널" 텍스트 명령어 대응 (관리자용)
    if message.content.strip() == "!상태패널":
        if isinstance(message.author, discord.Member) and message.author.guild_permissions.administrator:
            embed = discord.Embed(
                title="🏷️ 관전대기변경",
                description="아래 버튼을 클릭하여 닉네임 접두사를 변경할 수 있습니다.\n\n"
                            "📋 **사용 가능한 접두사**\n"
                            "• **관전** - `[관전] [닉네임]` 형태로 변경\n"
                            "• **대기** - `[대기] [닉네임]` 형태로 변경\n"
                            "• **원래대로** - 원래 닉네임으로 복원",
                color=discord.Color.blue()
            )
            await message.channel.send(embed=embed, view=StatusNicknameView())
        else:
            # DM이거나 관리자가 아닌 경우
            await message.channel.send("❌ 이 명령어는 관리자만 사용할 수 있습니다.", delete_after=5)
        return

    # "!팀나누기" 텍스트 명령어 대응 (모든 유저 가능)
    if message.content.strip().startswith("!팀나누기"):
        # 인원수 파싱 (예: !팀나누기 5 또는 !팀나누기)
        parts = message.content.strip().split()
        team_size = 5
        if len(parts) >= 2:
            try:
                team_size = int(parts[1])
            except ValueError:
                await message.reply("❌ 팀 인원수는 숫자여야 합니다. (예: `!팀나누기 5`)", delete_after=5)
                return
        
        if team_size <= 0:
            await message.reply("❌ 팀 인원수는 1명 이상이어야 합니다.", delete_after=5)
            return

        hub_channel_id = 1532691400230047805
        category_id = 1532692129569046559
        
        guild = message.guild
        hub_channel = guild.get_channel(hub_channel_id)
        category = guild.get_channel(category_id)
        
        if not hub_channel or not isinstance(hub_channel, discord.VoiceChannel):
            await message.reply("❌ 대상 대기방 음성 채널을 찾을 수 없습니다.", delete_after=5)
            return
            
        if not category or not isinstance(category, discord.CategoryChannel):
            await message.reply("❌ 대상 카테고리를 찾을 수 없습니다.", delete_after=5)
            return
            
        all_members = [m for m in hub_channel.members if not m.bot]
        active_members = [m for m in all_members if not m.display_name.startswith("[관전]")]
        spectator_members = [m for m in all_members if m.display_name.startswith("[관전]")]
        
        if not active_members:
            await message.reply("❌ 대상 대기방 채널에 플레이어가 없습니다. (관전자만 있는 경우 팀을 나눌 수 없습니다.)", delete_after=5)
            return
            
        random.shuffle(active_members)
        chunks = [active_members[i:i + team_size] for i in range(0, len(active_members), team_size)]
        
        status_msg = await message.reply("⏳ 팀 분배 및 채널 생성을 시작합니다...")
        
        import string
        created_channels = []
        failed_moves = 0
        success_moves = 0
        
        for chunk in chunks:
            existing_names = [c.name for c in category.voice_channels]
            new_name = None
            for char in string.ascii_uppercase:
                candidate = f"{char}팀"
                if candidate not in existing_names:
                    new_name = candidate
                    break
            if not new_name:
                i = 2
                while True:
                    candidate = f"A팀 {i}"
                    if candidate not in existing_names:
                        new_name = candidate
                        break
                    i += 1
            
            try:
                new_channel = await guild.create_voice_channel(name=new_name, category=category)
                created_channels.append(new_channel)
                for member in chunk:
                    try:
                        await member.move_to(new_channel)
                        success_moves += 1
                    except Exception:
                        failed_moves += 1
            except Exception as e:
                await status_msg.edit(content=f"❌ 오류가 발생했습니다: {e}")
                return
        
        # 관전자 이동 처리 (생성된 채널 중 랜덤 배정)
        if spectator_members and created_channels:
            for spectator in spectator_members:
                target_channel = random.choice(created_channels)
                try:
                    await spectator.move_to(target_channel)
                    success_moves += 1
                except Exception:
                    failed_moves += 1

        result_msg = f"✅ 플레이어 {len(active_members)}명을 {team_size}명씩 나누어 {len(chunks)}개의 팀 채널을 생성했습니다.\n"
        if spectator_members:
            result_msg += f"👁️ 관전자 {len(spectator_members)}명도 각 팀 채널에 랜덤하게 배정되었습니다.\n"
        result_msg += f"🔊 생성된 채널: {', '.join([c.name for c in created_channels])}\n"
        result_msg += f"👥 이동 완료: {success_moves}명"
        if failed_moves > 0:
            result_msg += f" (실패: {failed_moves}명)"
            
        await status_msg.edit(content=result_msg)
        return

    # 지정한 채널 ID 확인
    if message.channel.id == 1497843456960364726:
        content = message.content
        # 공백을 모두 제거하여 띄어쓰기 오차 무시
        content_stripped = "".join(content.split())
        
        # 유연한 키워드 매칭 조건 설정
        has_nickname = "닉네임/나이" in content_stripped or "닉넴/나이" in content_stripped
        has_game = "주로하는게임" in content_stripped
        has_time = "플레이시간대" in content_stripped
        has_intro = "소개글" in content_stripped
        
        if has_nickname and has_game and has_time and has_intro:
            member = message.author
            if isinstance(member, discord.Member):
                # 제거할 역할 (ID: 1369712767631626313)
                target_role_id = 1369712767631626313

                # [닉네임변경] 닉네임 자동 파싱 및 변경 기능 (활성화됨 - 특정 역할 보유자만)
                nickname = None
                nick_changed = False
                
                # 대상 역할(제거할 역할)을 가지고 있는 경우에만 닉네임 변경 진행
                if any(r.id == target_role_id for r in member.roles):
                    lines = content.split("\n")
                    for i, line in enumerate(lines):
                        line_stripped = "".join(line.split())
                        if "닉네임/나이" in line_stripped or "닉넴/나이" in line_stripped:
                            parts = line.split(":", 1)
                            if len(parts) < 2:
                                parts = line.split("-", 1)
                            
                            value_part = ""
                            if len(parts) >= 2:
                                value_part = parts[1].strip()
                            
                            # 콜론/대시 뒤가 비어있다면, 다음 줄 중 비어있지 않은 첫 번째 줄을 탐색
                            if not value_part:
                                for j in range(i + 1, len(lines)):
                                    next_line = lines[j].strip()
                                    if next_line:
                                        value_part = next_line
                                        break
                                        
                            if value_part:
                                # 패턴 1: 슬래시(/) 구분 (예: 오퍼 / 90, 너구리/95)
                                match1 = re.match(r'^(.+?)\s*/\s*(\d+)\s*(?:세|살)?(?:\D.*)?$', value_part)
                                if match1:
                                    nickname = match1.group(1).strip()
                                else:
                                    # 패턴 2: 공백 구분 (예: 희얼 03)
                                    match2 = re.match(r'^(.+?)\s+(\d+)\s*(?:세|살)?(?:\D.*)?$', value_part)
                                    if match2:
                                        nickname = match2.group(1).strip()
                            break
                    
                    if nickname is None:
                        # 닉네임 파싱 실패 시 역할 지급 중단 및 안내 메시지 전송 후 종료
                        await message.add_reaction("❌")
                        msg = await message.reply(
                            "❌ 닉네임/나이 양식이 올바르지 않습니다.\n"
                            "양식(예: `닉네임/나이: 홍길동 / 20` 또는 `닉네임/나이: 홍길동 20`)에 맞추어 정확히 다시 작성해 주세요.",
                            mention_author=False
                        )
                        await asyncio.sleep(5)
                        await msg.delete()
                        return

                    # 현재 별명에서 대괄호([]) 및 내부 텍스트 패턴 추출하여 접두사로 유지
                    current_nick = member.nick
                    if current_nick:
                        match = re.match(r'^(\[.*?\])\s*', current_nick)
                        if match:
                            prefix = match.group(1)
                            nickname = f"{prefix} {nickname}"

                    try:
                        await member.edit(nick=nickname)
                        print(f"✅ 닉네임 자동 변경 완료: {member.name} -> {nickname}")
                        nick_changed = True
                    except discord.Forbidden:
                        print(f"❌ 권한 부족: {member.name}의 닉네임을 {nickname}(으)로 변경할 수 없습니다.")
                    except Exception as e:
                        print(f"❌ 닉네임 변경 오류: {e}")

                role_to_remove = message.guild.get_role(target_role_id)
                
                # 부여할 역할 (ID: 1497939431473287238)
                grant_role_id = 1497939431473287238
                role_to_grant = message.guild.get_role(grant_role_id)
                
                # 역할 부여 처리
                grant_success = False
                if role_to_grant:
                    try:
                        await member.add_roles(role_to_grant)
                        print(f"✅ 역할 부여 완료: {member.name}에게 '{role_to_grant.name}' 역할을 부여했습니다.")
                        grant_success = True
                    except discord.Forbidden:
                        print(f"❌ 권한 부족: '{role_to_grant.name}' 역할을 부여할 수 없습니다. 봇의 역할 서열을 올려주세요.")
                    except Exception as e:
                        print(f"❌ 역할 부여 오류: {e}")
                else:
                    print(f"❌ 오류: 역할 ID {grant_role_id}를 서버에서 찾을 수 없습니다.")

                # 역할 제거 처리
                if role_to_remove:
                    if role_to_remove in member.roles:
                        try:
                            await member.remove_roles(role_to_remove)
                            await message.add_reaction("✅")
                            # 안내 메시지 전송 후 5초 뒤 자동 삭제
                            # [닉네임변경] 닉네임 변경 알림 포함 메시지 (활성화됨)
                            extra_msg = f" 아울러 닉네임이 **{nickname}**(으)로 변경되었습니다." if nick_changed else ""
                            msg = await message.reply(f"✅ 양식 작성이 확인되어 **{role_to_remove.name}** 역할이 제거되고 **{role_to_grant.name if role_to_grant else '새로운'}** 역할이 부여되었습니다.{extra_msg}", mention_author=False)
                            await asyncio.sleep(5)
                            await msg.delete()
                        except discord.Forbidden:
                            print(f"❌ 권한 부족: '{role_to_remove.name}' 역할을 제거할 수 없습니다. 봇의 역할 서열을 올려주세요.")
                        except Exception as e:
                            print(f"❌ 역할 제거 오류: {e}")
                    else:
                        # 이미 역할이 없는 경우에도 확인 리액션은 달아줌
                        await message.add_reaction("✅")
                        if grant_success and role_to_grant:
                            # [닉네임변경] 닉네임 변경 알림 포함 메시지 (활성화됨)
                            extra_msg = f" 아울러 닉네임이 **{nickname}**(으)로 변경되었습니다." if nick_changed else ""
                            msg = await message.reply(f"✅ 양식 작성이 확인되어 **{role_to_grant.name}** 역할이 부여되었습니다.{extra_msg}", mention_author=False)
                            await asyncio.sleep(5)
                            await msg.delete()
                else:
                    print(f"❌ 오류: 역할 ID {target_role_id}를 서버에서 찾을 수 없습니다.")

    await bot.process_commands(message)

# 유저 퇴장 감지 이벤트
@bot.event
async def on_member_remove(member):
    # 유저가 서버를 나가면 DB에 기록 및 과거 기록 수집
    import json
    last_application = None
    last_messages_list = []
    
    try:
        # 1. 가입 신청 채널(1497843456960364726)에서 작성했던 가입 신청서 찾기
        app_channel_id = 1497843456960364726
        app_channel = member.guild.get_channel(app_channel_id)
        if app_channel and isinstance(app_channel, discord.TextChannel):
            try:
                # 최근 200개 글 중 해당 멤버가 쓴 가입 양식 탐색
                async for msg in app_channel.history(limit=200):
                    if msg.author.id == member.id:
                        content_stripped = "".join(msg.content.split())
                        if any(k in content_stripped for k in ["닉네임", "닉넴", "나이", "게임", "소개"]):
                            last_application = msg.content
                            break
            except Exception as e:
                print(f"❌ 퇴장 멤버 가입서 조회 중 오류: {e}")

        # 2. 다른 텍스트 채널에서 작성한 최근 메시지 5개 수집
        collected_msgs = []
        try:
            # 봇이 읽을 수 있고 퇴장한 멤버가 보았을 법한 주요 텍스트 채널 탐색 (속도 제한을 위해 최대 15개 채널만 확인)
            text_channels = [c for c in member.guild.text_channels 
                             if c.permissions_for(member.guild.me).read_message_history and c.id != app_channel_id]
            for channel in text_channels[:15]:
                try:
                    async for msg in channel.history(limit=50):
                        if msg.author.id == member.id and msg.content:
                            # 봇 명령어 등은 제외
                            if not msg.content.startswith(('!', '/', '$', '?')):
                                collected_msgs.append({
                                    "channel": channel.name,
                                    "content": msg.content,
                                    "created_at": msg.created_at.isoformat()
                                })
                except Exception:
                    continue
        except Exception as e:
            print(f"❌ 퇴장 멤버 대화 메시지 수집 중 오류: {e}")

        # 시간순 정렬 후 최근 5개 선택
        collected_msgs.sort(key=lambda x: x["created_at"], reverse=True)
        last_messages_list = collected_msgs[:5]

    except Exception as e:
        print(f"❌ 퇴장 기록 수집 중 일반 오류 발생: {e}")

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        p = "%s" if DATABASE_URL else "?"
        
        # left_users 테이블에 수집된 정보와 함께 저장 (ON CONFLICT 업데이트 대응)
        if DATABASE_URL:
            # PostgreSQL
            query = """
                INSERT INTO left_users (user_id, last_application, last_messages)
                VALUES (%s, %s, %s)
                ON CONFLICT (user_id) 
                DO UPDATE SET 
                    last_application = EXCLUDED.last_application,
                    last_messages = EXCLUDED.last_messages
            """
            cursor.execute(query, (member.id, last_application, json.dumps(last_messages_list, ensure_ascii=False)))
        else:
            # SQLite
            query = """
                INSERT INTO left_users (user_id, last_application, last_messages)
                VALUES (?, ?, ?)
                ON CONFLICT (user_id) 
                DO UPDATE SET 
                    last_application = excluded.last_application,
                    last_messages = excluded.last_messages
            """
            cursor.execute(query, (member.id, last_application, json.dumps(last_messages_list, ensure_ascii=False)))
            
        conn.commit()
        cursor.close()
        conn.close()
        print(f"📥 유저 퇴장 기록 및 메시지 수집 완료: {member.name} ({member.id})")
    except Exception as e:
        print(f"❌ 퇴장 기록 중 DB 처리 오류 발생: {e}")

# 유저 입장 감지 이벤트
@bot.event
async def on_member_join(member):
    # 최근 업데이트: 입장 시 1369712767631626313 부여 및 1497939431473287238 제거
    # 입장 시 역할 부여 (역할 ID: 1369712767631626313)
    target_role_id = 1369712767631626313
    role = member.guild.get_role(target_role_id)
    if role:
        try:
            await member.add_roles(role)
            print(f"✅ 역할 부여 완료: {member.name}에게 '{role.name}' 역할을 부여했습니다.")
        except discord.Forbidden:
            print(f"❌ 권한 부족: '{role.name}' 역할을 부여할 수 없습니다. 봇의 역할 서열을 올려주세요.")
        except Exception as e:
            print(f"❌ 역할 부여 중 오류 발생: {e}")
    else:
        print(f"❌ 오류: 역할 ID {target_role_id}를 서버에서 찾을 수 없습니다.")

    # 다른 자동화 시스템/봇에 의해 역할이 부여될 시간을 확보하기 위해 2초 대기 후 최신 멤버 정보 로드
    await asyncio.sleep(2)
    try:
        fresh_member = await member.guild.fetch_member(member.id)
    except Exception as fetch_err:
        print(f"❌ 최신 멤버 정보 로드 실패: {fetch_err}")
        fresh_member = member

    # 입장 시 역할 제거 (역할 ID: 1497939431473287238)
    remove_role_id = 1497939431473287238
    role_to_remove = member.guild.get_role(remove_role_id)
    if role_to_remove:
        try:
            await fresh_member.remove_roles(role_to_remove)
            print(f"✅ 역할 제거 완료: {fresh_member.name}에게서 '{role_to_remove.name}' 역할을 제거했습니다.")
        except discord.Forbidden:
            print(f"❌ 권한 부족: '{role_to_remove.name}' 역할을 제거할 수 없습니다. 봇의 역할 서열을 올려주세요.")
        except Exception as e:
            print(f"❌ 역할 제거 중 오류 발생: {e}")
    else:
        print(f"❌ 오류: 역할 ID {remove_role_id}를 서버에서 찾을 수 없습니다.")
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        p = "%s" if DATABASE_URL else "?"
        cursor.execute(f"SELECT user_id, last_application, last_messages FROM left_users WHERE user_id = {p}", (member.id,))
        row = cursor.fetchone()
        
        if row:
            # 이전에 나갔던 기록이 있는 유저가 다시 들어온 경우
            user_id = row[0]
            last_app = row[1]
            last_msgs_json = row[2]
            
            channel_id = 1498300372479901817
            channel = bot.get_channel(channel_id)
            if channel:
                try:
                    embed = discord.Embed(
                        title="👤 재입장 감지",
                        description=f"<@{member.id}> (ID: {member.id}) 님이 서버에 다시 입장하셨습니다.",
                        color=discord.Color.green(),
                        timestamp=datetime.datetime.now()
                    )
                    
                    if last_app:
                        # 글자수 제한 1024자 고려
                        app_display = last_app[:1024] if len(last_app) <= 1024 else last_app[:1021] + "..."
                        embed.add_field(name="📝 과거 작성한 가입 신청서", value=f"```\n{app_display}\n```", inline=False)
                        
                    if last_msgs_json:
                        try:
                            last_msgs = json.loads(last_msgs_json)
                            if last_msgs:
                                msg_lines = []
                                for m in last_msgs:
                                    # 시간 파싱
                                    try:
                                        dt = datetime.datetime.fromisoformat(m["created_at"])
                                        time_str = dt.strftime("%m/%d %H:%M")
                                    except Exception:
                                        time_str = "시간 정보 없음"
                                    content_display = m["content"][:100] if len(m["content"]) <= 100 else m["content"][:97] + "..."
                                    msg_lines.append(f"[{time_str}] #{m['channel']}: {content_display}")
                                
                                embed.add_field(name="💬 마지막으로 작성한 메시지 (최근 5개)", value="\n".join(msg_lines), inline=False)
                        except Exception as json_err:
                            print(f"❌ 메시지 JSON 파싱 오류: {json_err}")
                            
                    await channel.send(embed=embed)
                except Exception as send_err:
                    print(f"❌ 메시지 전송 실패: {send_err}")
            else:
                print(f"❌ 알림 채널({channel_id})을 찾을 수 없습니다.")
                
            # 기록에서 삭제
            cursor.execute(f"DELETE FROM left_users WHERE user_id = {p}", (member.id,))
            conn.commit()
            print(f"📤 재입장 확인 후 DB 기록 삭제: {member.name} ({member.id})")
            
        cursor.close()
        conn.close()
    except Exception as e:
        print(f"❌ 입장 확인 중 오류 발생: {e}")

# 동적 음성 채널 생성 이벤트
@bot.event
async def on_voice_state_update(member, before, after):
    # 0. 음성 채널 이용 시간 기록
    if before.channel != after.channel:
        # 음성 채널에 새로 입장한 경우
        if before.channel is None and after.channel is not None:
            if not member.bot:
                active_sessions[member.id] = time.time()
                print(f"🎙️ 음성 채널 입장 감지: {member.name} ({member.id})")
        # 음성 채널에서 완전히 퇴장한 경우
        elif before.channel is not None and after.channel is None:
            if not member.bot:
                join_time = active_sessions.pop(member.id, None)
                if join_time:
                    duration = int(time.time() - join_time)
                    if duration > 0:
                        add_voice_time(member.id, get_current_date(), duration)
                        print(f"🎙️ 음성 채널 퇴장 감지: {member.name} ({member.id}) - 이용 시간: {duration}초 추가")

    # 다중 허브 채널 및 대상 설정
    CONFIGS = {
        1511038705932963991: {
            "category_id": 1357930212146286644,
            "channel_name": "⚡・『 메인게임 』"
        },
        1511037391765504130: {
            "category_id": 1427312936098992262,
            "channel_name": "💫・『 종합게임 』"
        },
        1510992054174351510: {
            "category_id": 1511040771241935069,
            "channel_name": "🫧・『 싱글게임 』"
        }
    }

    # 1. 허브 채널 입장 감지 및 채널 생성
    if after.channel and after.channel.id in CONFIGS:
        config = CONFIGS[after.channel.id]
        guild = member.guild
        category = guild.get_channel(config["category_id"])
        
        if category and isinstance(category, discord.CategoryChannel):
            try:
                # 채널 이름 결정
                channel_name = config["channel_name"]
                if config.get("use_alphabet"):
                    existing_names = [c.name for c in category.voice_channels]
                    import string
                    new_name = None
                    for char in string.ascii_uppercase:
                        candidate = f"{char}팀"
                        if candidate not in existing_names:
                            new_name = candidate
                            break
                    if not new_name:
                        # A-Z까지 전부 꽉 찬 경우 예외적으로 A팀 2, A팀 3 ... 검색
                        i = 2
                        while True:
                            candidate = f"A팀 {i}"
                            if candidate not in existing_names:
                                new_name = candidate
                                break
                            i += 1
                    channel_name = new_name

                # 지정된 카테고리 하위에 새 음성 채널 생성
                new_channel = await guild.create_voice_channel(
                    name=channel_name,
                    category=category
                )
                print(f"🔊 새 음성 채널 생성 완료: '{channel_name}' (ID: {new_channel.id})")
                
                # 유저를 생성된 채널로 이동
                await member.move_to(new_channel)
                print(f"➡️ {member.name} 님을 '{channel_name}' 채널로 이동시켰습니다.")
            except discord.Forbidden:
                print("❌ 권한 부족: 채널 생성 또는 멤버 이동 권한이 없습니다.")
            except Exception as e:
                print(f"❌ 음성 채널 생성/이동 중 오류 발생: {e}")
        else:
            print(f"❌ 오류: 카테고리 ID {config['category_id']}를 찾을 수 없거나 올바른 카테고리가 아닙니다.")

    # 2. 유저 퇴장 감지 및 빈 임시 채널 삭제
    if before.channel and before.channel != after.channel:
        # 퇴장 감지할 카테고리 ID 목록 수집
        target_categories = [config["category_id"] for config in CONFIGS.values()]
        # 팀 나누기 카테고리 ID(1532692129569046559)도 자동 삭제 대상 카테고리에 포함시킴
        team_category_id = 1532692129569046559
        if team_category_id not in target_categories:
            target_categories.append(team_category_id)
            
        # 허브 채널 ID 목록 수집 (허브 채널 자체는 삭제 방지)
        hub_channel_ids = list(CONFIGS.keys())
        # 팀 나누기 허브 채널 ID도 삭제 방지 목록에 포함시킴
        team_hub_id = 1532691400230047805
        if team_hub_id not in hub_channel_ids:
            hub_channel_ids.append(team_hub_id)

        if before.channel.category and before.channel.category.id in target_categories:
            # 허브 채널이 아니고 빈 채널이면 삭제 (이름 변경 지원)
            if before.channel.id not in hub_channel_ids and len(before.channel.members) == 0:
                try:
                    await before.channel.delete()
                    print(f"🗑️ 빈 음성 채널 삭제 완료: {before.channel.name} (ID: {before.channel.id})")
                except discord.Forbidden:
                    print("❌ 권한 부족: 채널을 삭제할 수 없습니다.")
                except Exception as e:
                    print(f"❌ 음성 채널 삭제 중 오류 발생: {e}")



# Run Flask server and start Discord bot
if __name__ == "__main__":
    keep_alive()
    token = os.getenv("TOKEN")
    if token:
        bot.run(token)
    else:
        print("⚠️ TOKEN 환경변수가 설정되지 않았습니다.")